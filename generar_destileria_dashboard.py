#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_destileria_dashboard.py
Genera destileria_dashboard.html con datos de:
  temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final

Uso local:
  python -X utf8 generar_destileria_dashboard.py

Uso Cloud Run (sube automáticamente a GCS):
  python3 generar_destileria_dashboard.py \
      --output /tmp/destileria_dashboard.html \
      --gcs-bucket temple-bar-dashboard-cache
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
import google.auth
from google.cloud import bigquery

PROJECT   = "temple-brewery"
TABLE     = "temple-brewery.Destileria.vw_ventas_con_cluster"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE  = os.path.join(SCRIPT_DIR, "templates", "destileria.html")
OUTPUT_DEFAULT  = os.path.join(SCRIPT_DIR, "destileria_dashboard.html")
OBJ_JSON_FILE   = os.path.join(SCRIPT_DIR, "objetivos_destileria.json")

SHEET_OBJ_ID = os.environ.get("DEST_OBJ_DRIVE_ID", "1curY4eZKp6WZ_r2p8W9sglsdY3UUzirx")
OBJ_GCS_BLOB = "objetivos_destileria.json"   # cache persistente en GCS

SA_KEY = os.path.join(SCRIPT_DIR, "temple-bar-439715-da51b292ce5d.json")

GCP_SCOPES = [
    "https://www.googleapis.com/auth/bigquery",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]

# Mapa: nombre de producto en el Sheet (MAYÚSCULAS) → familia de classify_familia
PRODUCT_FAMILIA_MAP = {
    "GIN BOSQUE ALTA MONTAÑA BOTELLA 500 ML":          "bosque_alta_montana",
    "GIN BOSQUE ALTA MONTAÑA BOTELLA 750 ML":          "bosque_alta_montana",
    "GIN BOSQUE ALTA MONTAÑA MINIATURA BOTELLA 50 ML": "bosque_alta_montana_mini",
    "GIN BOSQUE ALTA MONTANA BOTELLA 500 ML":          "bosque_alta_montana",
    "GIN BOSQUE ALTA MONTANA BOTELLA 750 ML":          "bosque_alta_montana",
    "GIN BOSQUE ALTA MONTANA MINIATURA BOTELLA 50 ML": "bosque_alta_montana_mini",
    "GIN BOSQUE NATIVO BOTELLA 500 ML":                "bosque_nativo",
    "GIN BOSQUE NATIVO BOTELLA 750 ML":                "bosque_nativo",
    "GIN BOSQUE NATIVO MINIATURA BOTELLA 50 ML":       "bosque_nativo_mini",
    "GIN BOSQUE REFUGIOS BOTELLA 500 ML":              "bosque_refugios",
    "GIN BOSQUE REFUGIOS BOTELLA 750 ML":              "bosque_refugios",
    "VERMU FERIADO ROJO 750 ML":                       "feriado_rojo",
    "VERMÚ FERIADO ROJO 750 ML":                       "feriado_rojo",
    "VERMU FERIADO ROJO BARRIL 20 LTS":                "feriado_barril_20",
    "VERMU FERIADO ROJO BARRIL 20 LT":                 "feriado_barril_20",
    "VERMU FERIADO ROJO BARRIL 20 L":                  "feriado_barril_20",
    "VERMU FERIADO ROSADO 750 ML":                     "feriado_rosado",
    "VERMÚ FERIADO ROSADO 750 ML":                     "feriado_rosado",
    "WOLF IPA":                                        "lata_wolf",
    "WOLF IPA (LATA)":                                 "lata_wolf",
    "SCOTTISH":                                        "lata_scottish",
    "SCOTTISH (LATA)":                                 "lata_scottish",
    "WOLF IPA 0%":                                     "lata_wolf0",
    "WOLF IPA 0% ALC":                                 "lata_wolf0",
    "WOLF IPA 0% (LATA)":                              "lata_wolf0",
    "INDIE GOLDEN":                                    "lata_golden",
    "INDIE GOLDEN (LATA)":                             "lata_golden",
    "GOLDEN LAGER MUNDIAL":                            "lata_golden",
    "FLOW APA (LATA)":                                 "lata_otras",
    "BLACK SOUL STOUT (LATA)":                         "lata_otras",
    "COSMICA (LATA)":                                  "lata_otras",
}

SKIP_CLUSTER_PATTERNS = ["OBJ TOTAL", "BXQ", "SUPERMERCADOS"]


# ---------------------------------------------------------------------------
# GCS upload
# ---------------------------------------------------------------------------

def upload_to_gcs(local_path, bucket_name, blob_name="destileria_dashboard.html", html_content=None):
    """Sube el HTML generado a GCS con cache-control adecuado."""
    from google.cloud import storage
    print(f"\nUploading to GCS: gs://{bucket_name}/{blob_name} ...")
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    if html_content is not None:
        # Upload directo desde memoria — evita race condition con Drive Sync
        content_bytes = html_content.encode("utf-8")
        file_size = len(content_bytes)
        if file_size < 1024:
            raise RuntimeError(f"HTML demasiado pequeño ({file_size} bytes) — posible archivo corrupto, abortando upload")
        blob.upload_from_string(content_bytes, content_type="text/html; charset=utf-8")
    else:
        file_size = os.path.getsize(local_path)
        if file_size < 1024:
            raise RuntimeError(f"HTML demasiado pequeño ({file_size} bytes) — posible archivo corrupto, abortando upload")
        blob.upload_from_filename(local_path, content_type="text/html; charset=utf-8")
    blob.cache_control = "no-cache, no-store, must-revalidate"
    blob.patch()
    blob.reload()
    if blob.cache_control != "no-cache, no-store, must-revalidate":
        print(f"  WARN: cache_control no aplicado correctamente (valor: {blob.cache_control})", file=__import__('sys').stderr)
    public_url = f"https://storage.googleapis.com/{bucket_name}/{blob_name}"
    print(f"  OK: {public_url} ({file_size // 1024} KB)")
    return public_url


# ---------------------------------------------------------------------------
# Firestore — objetivos (fuente primaria)
# ---------------------------------------------------------------------------

def load_objectives_from_firestore(db):
    """
    Carga objetivos desde la colección 'objetivos_destileria' en Firestore.
    Cada documento tiene la forma:
        {marca, dimension, nombre, valores: [12 ints], updated_at, updated_by}
    Devuelve dict {marca: {dimension: {nombre: [12 valores]}}} o {} si vacío/error.
    """
    try:
        result = {}
        docs = list(db.collection("objetivos_destileria").stream())
        if not docs:
            return {}
        for doc in docs:
            d = doc.to_dict()
            marca     = d.get("marca")
            dimension = d.get("dimension")
            nombre    = d.get("nombre")
            valores   = d.get("valores")
            if not (marca and dimension and nombre and isinstance(valores, list) and len(valores) == 12):
                print(
                    f"WARN: Firestore doc '{doc.id}' ignorado — campos incompletos o valores inválidos",
                    file=__import__('sys').stderr,
                )
                continue
            result.setdefault(marca, {}).setdefault(dimension, {})[nombre] = valores
        return result
    except Exception as _fs_err:
        print(f"WARN: Firestore objetivos falló: {_fs_err}", file=__import__('sys').stderr)
        return {}


# ---------------------------------------------------------------------------
# GCS cache de objetivos (persiste entre runs de Cloud Run)
# ---------------------------------------------------------------------------

def save_objectives_to_gcs(obj_data, bucket_name, blob_name=OBJ_GCS_BLOB):
    """Guarda el JSON de objetivos en GCS como backup persistente entre runs."""
    from google.cloud import storage
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    content = json.dumps(obj_data, ensure_ascii=False, indent=2)
    blob.upload_from_string(content.encode("utf-8"), content_type="application/json; charset=utf-8")
    blob.cache_control = "no-cache, no-store, must-revalidate"
    blob.patch()
    print(f"  OK: objetivos guardados en gs://{bucket_name}/{blob_name}")


def load_objectives_from_gcs(bucket_name, blob_name=OBJ_GCS_BLOB):
    """Carga el JSON de objetivos desde GCS. Lanza excepción si no existe."""
    from google.cloud import storage
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    content = blob.download_as_text(encoding="utf-8")
    return json.loads(content)


# ---------------------------------------------------------------------------
# Clasificación de productos
# ---------------------------------------------------------------------------

def classify_familia(producto, envase):
    """Devuelve la familia del producto para agrupar en el dashboard."""
    p = str(producto or "").upper()
    e = str(envase  or "").upper()

    # Barriles (Feriado) — detectar por envase primero, distinguir tamaño
    if "BARRIL" in e:
        if "50" in e or "50" in p: return "feriado_barril_50"
        return "feriado_barril_20"

    # Bosque — miniaturas antes que los genéricos
    if "BOSQUE NATIVO"   in p and "MINI" in p: return "bosque_nativo_mini"
    if "BOSQUE" in p and "ALTA MONTA" in p and "MINI" in p: return "bosque_alta_montana_mini"

    # Bosque
    if "BOSQUE NATIVO"       in p: return "bosque_nativo"
    if "BOSQUE ALTA MONTA"   in p: return "bosque_alta_montana"
    if "BOSQUE REFUGIOS"     in p: return "bosque_refugios"
    if "GIN RIDER"           in p: return "bosque_otro"
    if re.search(r'\bGIN\b', p): return "bosque_otro"  # palabra completa — evita falso positivo con "GINEBRA"

    # Feriado
    if "FERIADO ROSADO"      in p: return "feriado_rosado"
    if "FERIADO ROJO"        in p: return "feriado_rojo"
    if "VERMU"               in p: return "feriado_rojo"
    if "VERM\u00da"          in p: return "feriado_rojo"

    # Cervezas en lata — detectar por envase
    if "LATA" in e:
        if "WOLF IPA" in p and "0%" in p:           return "lata_wolf0"
        if "WOLF IPA" in p:                          return "lata_wolf"
        if "SCOTTISH" in p:                          return "lata_scottish"
        if "GOLDEN" in p or "INDIE" in p:            return "lata_golden"
        if "APA" in p or "IPL" in p or "IPA" in p:  return "lata_otras"
        return "lata_otras"

    # Todo lo demás → Merch (botellas de terceros, complementarios, etc.)
    return "merch"


# ---------------------------------------------------------------------------
# Objetivos — parsing Google Sheets
# ---------------------------------------------------------------------------

def parse_obj_num(s):
    """Parsea número del Sheet.
    - '1,500' o '1.500' (miles europeo, exactamente 3 decimales) → 1500
    - '29,43' (decimal con coma) → 29
    - '1775.114372' o '1499.5' (float decimal) → redondeado al entero más cercano
    """
    s = str(s).strip().replace("\xa0", "").replace("$", "").replace(" ", "")
    if not s or s in ("-", "—"):
        return 0
    # Coma decimal: "29,43" (1-3 dígitos + coma + 1-2 decimales)
    if re.match(r"^\d{1,3},\d{1,2}$", s):
        return round(float(s.replace(",", ".")))
    # Miles europeo con punto: "3.274" o "1.500" (≤3 dígitos antes, exactamente 3 después)
    # Si el resultado supera 50.000 es un decimal mal interpretado (ej: "328.545" → 329 L)
    if re.match(r"^\d{1,3}\.\d{3}$", s):
        as_european = int(s.replace(".", ""))
        if as_european > 50_000:
            return round(float(s))
        return as_european
    # Float decimal o entero: "1775.114372", "1499.5", "2349.0", "1882"
    try:
        return round(float(s.replace(",", "")))
    except Exception:
        return 0


def _month_cols(header_row):
    """Lista de 12 índices de columna (0-based) para ene–dic."""
    mmap = {
        "enero": 0, "ene": 0, "feb": 1, "febrero": 1, "mar": 2, "marzo": 2,
        "abr": 3, "abril": 3, "may": 4, "mayo": 4, "jun": 5, "junio": 5,
        "jul": 6, "julio": 6, "ago": 7, "agosto": 7,
        "sep": 8, "sept": 8, "septiembre": 8,
        "oct": 9, "octubre": 9, "nov": 10, "noviembre": 10,
        "dic": 11, "diciembre": 11,
    }
    found = {}
    for i, c in enumerate(header_row):
        k = str(c).strip().lower()
        if k in mmap and mmap[k] not in found:
            found[mmap[k]] = i
    return [found.get(m, -1) for m in range(12)]


def _is_header(row):
    """True si la fila contiene ≥4 nombres de meses."""
    mkeys = {"enero","feb","febrero","mar","marzo","abr","abril","may","mayo",
             "jun","junio","jul","julio","ago","agosto","sep","sept","oct","nov","dic"}
    return sum(1 for c in row if str(c).strip().lower() in mkeys) >= 4


def _detect_brand(row):
    """Detecta marcador de sección. Devuelve 'feriado' | 'cerveza' | None.
    Solo activa si la primera celda es corta (≤20 chars), para no confundir
    nombres de productos (ej: 'VERMU FERIADO ROJO 750 ML') con encabezados."""
    if not row:
        return None
    first = str(row[0]).strip()
    if len(first) > 20:
        return None
    txt = " ".join(str(c) for c in row[:4]).upper()
    if "FERIADO" in txt and "GIN" not in txt and "BOSQUE" not in txt:
        return "feriado"
    if "CERVEZA" in txt:
        return "cerveza"
    return None


def _norm(s):
    """Normaliza string para comparar sin tildes."""
    return (s.upper()
              .replace("Á", "A").replace("É", "E").replace("Í", "I")
              .replace("Ó", "O").replace("Ú", "U").replace("Ñ", "N"))


def _merge(dest, key, vals):
    """Agrega vals a dest[key] (suma si ya existe)."""
    dest[key] = [dest[key][m] + vals[m] for m in range(12)] if key in dest else vals


def _parse_product_sheet(rows, result):
    """Parsea hoja 1 (objetivos por etiqueta/producto).
    `seen` no se resetea por marca: sk=(brand,key) ya identifica la marca,
    y debe persistir para descartar el segundo bloque "sensibilizado" que
    repite el encabezado de la misma marca más abajo en la hoja."""
    brand, month_cols, seen = "bosque", None, set()
    for row in rows:
        if not row:
            continue
        nb = _detect_brand(row)
        if nb:
            brand, month_cols = nb, None
            continue
        if _is_header(row):
            month_cols = _month_cols(row)
            continue
        if month_cols is None:
            continue
        name = _norm(str(row[0]).strip())
        if not name:
            continue
        if "VENTAS TOTALES" in name:
            key = "_TOTAL"
        else:
            key = PRODUCT_FAMILIA_MAP.get(name)
            if key is None:
                for pn, pk in PRODUCT_FAMILIA_MAP.items():
                    if _norm(pn) == name:
                        key = pk
                        break
        if key is None:
            continue
        sk = (brand, key)
        if sk in seen:       # skip segundo bloque "sensibilizado"
            continue
        seen.add(sk)
        vals = [parse_obj_num(row[ci]) if ci >= 0 and ci < len(row) else 0
                for ci in month_cols]
        _merge(result[brand]["product"], key, vals)


def _parse_cluster_sheet(rows, result):
    """Parsea hoja 2 (objetivos por cluster).
    `seen` no se resetea por marca: sk=(brand,key) ya identifica la marca,
    y debe persistir para descartar el segundo bloque "sensibilizado" que
    repite el encabezado de la misma marca más abajo en la hoja."""
    brand, month_cols, seen = "bosque", None, set()
    for row in rows:
        if not row:
            continue
        nb = _detect_brand(row)
        if nb:
            brand, month_cols = nb, None
            continue
        if _is_header(row):
            month_cols = _month_cols(row)
            continue
        if month_cols is None:
            continue
        name = str(row[0]).strip()
        if not name:
            continue
        nu = name.upper()
        if "VENTAS TOTALES" in nu:
            key = "_TOTAL"
        else:
            if any(p in nu for p in SKIP_CLUSTER_PATTERNS):
                continue
            first = str(row[1]).strip() if len(row) > 1 else ""
            if first == "-":
                continue
            key = name
        sk = (brand, key)
        if sk in seen:
            continue
        seen.add(sk)
        vals = [parse_obj_num(row[ci]) if ci >= 0 and ci < len(row) else 0
                for ci in month_cols]
        if key != "_TOTAL" and not any(v > 0 for v in vals):
            continue
        _merge(result[brand]["cluster"], key, vals)


def fetch_objectives(creds):
    """
    Descarga el archivo Excel de objetivos desde Google Drive y lo parsea.
    Funciona con archivos .xlsx (no nativos de Google Sheets).
    Devuelve dict: {brand: {"product": {...}, "cluster": {...}}}
    Cada valor es lista de 12 ints (ene–dic).
    """
    import io
    import openpyxl
    import googleapiclient.discovery as disc
    from googleapiclient.http import MediaIoBaseDownload

    # Intentar SA key primero (tiene scope drive), luego ADC como fallback
    _drive_creds_list = []
    if os.path.exists(SA_KEY):
        from google.oauth2 import service_account as _sa
        _drive_creds_list.append(("SA key", _sa.Credentials.from_service_account_file(
            SA_KEY, scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )))
    _drive_creds_list.append(("ADC", creds))
    drive_svc = None
    for _label, _dc in _drive_creds_list:
        try:
            _svc = disc.build("drive", "v3", credentials=_dc, cache_discovery=False)
            _svc.files().get(fileId=SHEET_OBJ_ID, fields="id").execute()
            drive_svc = _svc
            print(f"  Drive auth OK ({_label})")
            break
        except Exception as _auth_err:
            print(f"  Drive auth fallida con {_label}: {_auth_err!r}")
    if drive_svc is None:
        raise RuntimeError("No se pudo autenticar con Drive (ni SA key ni ADC)")

    print(f"  Descargando Excel (ID: {SHEET_OBJ_ID})...")
    for _attempt in range(1, 4):
        try:
            request = drive_svc.files().get_media(fileId=SHEET_OBJ_ID)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            fh.seek(0)
            break
        except Exception as _dl_err:
            if _attempt == 3:
                raise
            import time as _time
            print(f"  Intento {_attempt} fallido ({_dl_err!r}), reintentando...", file=__import__('sys').stderr)
            _time.sleep(5)

    wb = openpyxl.load_workbook(fh, data_only=True)
    names = wb.sheetnames
    PRODUCT_SHEET = "Rolling - Mensual x Etiqueta"
    CLUSTER_SHEET = "Rolling - Mensual x Cluster"
    if PRODUCT_SHEET not in names or CLUSTER_SHEET not in names:
        raise ValueError(f"Hojas esperadas no encontradas. Disponibles: {names}")
    t1 = PRODUCT_SHEET
    t2 = CLUSTER_SHEET
    print(f"  Hoja productos: {t1}")
    print(f"  Hoja clusters:  {t2}")

    def sheet_to_rows(sheet_name):
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        return rows

    result = {b: {"product": {}, "cluster": {}} for b in ("bosque", "feriado", "cerveza")}
    _parse_product_sheet(sheet_to_rows(t1), result)
    if t2:
        _parse_cluster_sheet(sheet_to_rows(t2), result)

    # Cerveza: objetivos en latas (473 ml c/u) → convertir a litros para comparar con BQ
    # Excepción: product._TOTAL ("VENTAS TOTALES EN LTS" en hoja "Rolling - Mensual x Etiqueta")
    # ya viene cargado directamente en litros, no en latas — no convertir esa fila.
    L_POR_LATA = 0.473
    for section in ("product", "cluster"):
        for key, arr in result["cerveza"][section].items():
            if section == "product" and key == "_TOTAL":
                continue
            result["cerveza"][section][key] = [round(v * L_POR_LATA, 1) for v in arr]

    return result


# ---------------------------------------------------------------------------
# Validación de objetivos
# ---------------------------------------------------------------------------

# Rangos razonables de litros mensuales por marca (mínimo, máximo)
_OBJ_RANGES = {
    "bosque":  (200,   6_000),
    "feriado": (200,   5_000),
    "cerveza": (500,  30_000),   # objetivos en litros (convertidos desde latas × 0.473)
}
# Clusters que deben estar sí o sí
_OBJ_MIN_CLUSTERS = {
    "bosque":  {"Cadena Grupo Temple", "Distribuidor"},
    "feriado": {"Cadena Grupo Temple", "Distribuidor"},
    "cerveza": {"Grupo Temple", "Distribuidores"},   # nombres actuales en el Excel
}
# Productos que deben estar sí o sí (cerveza no tiene breakdown obligatorio)
_OBJ_MIN_PRODUCTS = {
    "bosque":  {"bosque_nativo", "bosque_alta_montana"},
    "feriado": {"feriado_rojo",  "feriado_rosado"},
    "cerveza": set(),
}
_OBJ_CHANGE_THRESHOLD = 0.25   # >25% de cambio vs caché anterior = advertencia


def validate_objectives(obj, prev_obj=None):
    """
    Valida estructura y valores de objetivos.
    Retorna (errores: list[str], advertencias: list[str]).
    Errores = datos inválidos/faltantes.
    Advertencias = cambios grandes vs caché anterior (puede ser legítimo).
    """
    errors, warns = [], []
    for brand in ("bosque", "feriado", "cerveza"):
        b = obj.get(brand, {})
        lo, hi = _OBJ_RANGES[brand]

        total = (b.get("cluster", {}).get("_TOTAL")
                 or b.get("product", {}).get("_TOTAL"))
        if not total:
            errors.append(f"[{brand}] _TOTAL ausente — parsing probablemente falló")
            continue
        if len(total) != 12:
            errors.append(f"[{brand}] _TOTAL tiene {len(total)} valores (esperados 12)")

        # Ceros en los primeros 6 meses (período con datos completos)
        zeros = [i + 1 for i, v in enumerate(total[:6]) if v == 0]
        if zeros:
            errors.append(f"[{brand}] _TOTAL con ceros en meses {zeros}")

        # Valores fuera de rango razonable
        for i, v in enumerate(total):
            if v > 0 and not (lo <= v <= hi):
                errors.append(
                    f"[{brand}] mes {i+1} = {v:,.0f} L fuera del rango [{lo:,}–{hi:,}]"
                )

        # Clusters mínimos (comparación case-insensitive + strip)
        cl_keys = set(b.get("cluster", {}).keys()) - {"_TOTAL"}
        cl_keys_norm = {k.strip().lower() for k in cl_keys}
        for exp in _OBJ_MIN_CLUSTERS.get(brand, set()):
            if exp.strip().lower() not in cl_keys_norm:
                errors.append(f"[{brand}] cluster obligatorio faltante: '{exp}' (disponibles: {sorted(cl_keys)})")

        # Productos mínimos
        prod_keys = set(b.get("product", {}).keys()) - {"_TOTAL"}
        for exp in _OBJ_MIN_PRODUCTS.get(brand, set()):
            if exp not in prod_keys:
                errors.append(f"[{brand}] producto obligatorio faltante: '{exp}'")

        # Cambio grande vs caché anterior
        if prev_obj:
            prev_b = prev_obj.get(brand, {})
            prev_total = (prev_b.get("cluster", {}).get("_TOTAL")
                          or prev_b.get("product", {}).get("_TOTAL"))
            if prev_total:
                for i in range(min(len(total), len(prev_total))):
                    pv, nv = prev_total[i], total[i]
                    if pv > 0 and abs(nv - pv) / pv > _OBJ_CHANGE_THRESHOLD:
                        warns.append(
                            f"[{brand}] mes {i+1}: {pv:,.0f} → {nv:,.0f} L "
                            f"({(nv - pv) / pv * 100:+.1f}%) — verificar si es cambio intencional"
                        )
    return errors, warns


# ---------------------------------------------------------------------------
# Fetch BQ
# ---------------------------------------------------------------------------

def fetch_rows(client):
    query = f"""
    WITH base AS (
      -- Deduplica filas idénticas ignorando Clusterizacion (que puede ser NULL en algunas copias)
      SELECT DISTINCT
        FechaPedido, NombreDeFantasia, GrupoCliente,
        Producto, Envase, Tipo, CantEnvases, Litros, Total_
      FROM `{TABLE}`
      WHERE FechaPedido IS NOT NULL
        AND FechaPedido < '2026-07-01'   -- datos históricos solo hasta el corte; julio en adelante viene de Contabilium
    ),
    cl_latest AS (
      -- Cluster más reciente no-null por cliente
      SELECT DISTINCT NombreDeFantasia, Clusterizacion
      FROM `{TABLE}`
      WHERE Clusterizacion IS NOT NULL
      QUALIFY ROW_NUMBER() OVER (
        PARTITION BY NombreDeFantasia ORDER BY FechaPedido DESC
      ) = 1
    )
    SELECT
      FORMAT_DATE('%Y-%m-%d', b.FechaPedido)              AS f,
      COALESCE(c.Clusterizacion, 'Sin Cluster')            AS cl,
      COALESCE(b.NombreDeFantasia, 'Sin nombre')           AS nd,
      COALESCE(b.GrupoCliente, '')                         AS gc,
      COALESCE(b.Producto, '')                             AS pr,
      COALESCE(b.Envase, '')                               AS en,
      COALESCE(b.Tipo, '')                                 AS ti,
      COALESCE(b.CantEnvases, 0)                           AS ce,
      ROUND(COALESCE(b.Litros,  0.0), 3)                   AS li,
      ROUND(COALESCE(b.Total_,  0.0), 0)                   AS to_
    FROM base b
    LEFT JOIN cl_latest c ON b.NombreDeFantasia = c.NombreDeFantasia
    ORDER BY b.FechaPedido
    """
    return list(client.query(query).result(timeout=120))


# Mapeo de nombre Contabillium → cluster correcto (clientes sin match en cluster_map)
_CBL_CLUSTER_MAP: dict[str, str] = {
    "TEMPLE HOLLYWOOD NUEVO":       "Cadena Grupo Temple",
    "TEMPLE MASCHWITZ":             "Cadena Grupo Temple",
    "TEMPLE PALERMO NUEVO":         "Cadena Grupo Temple",
    "TEMPLE PASEO LA PLAZA":        "Cadena Grupo Temple",
    "TEMPLE RECOLETA SRL":          "Cadena Grupo Temple",
    "TEMPLE RIO GALLEGOS":          "Cadena Grupo Temple",
    "TEMPLE SALTA":                 "Cadena Grupo Temple",
    "TEMPLE SOHO":                  "Cadena Grupo Temple",
    "LOS TEMPLOS CABALLITO S.R.L.": "Cadena Grupo Temple",
    "REBELION":                     "Cadena Grupo Temple",
    "BARRA PATIO DE LOS LECHEROS":  "Bar/Resto",
}


def fetch_rows_contabilium(creds):
    """Trae ventas de Contabilium (temple-bar-439715) desde 2026-07-01.
    Devuelve SimpleNamespace con los mismos campos que fetch_rows()."""
    import types
    import google.cloud.bigquery as _bq
    client = _bq.Client(project="temple-bar-439715", credentials=creds)
    query = """
    WITH cluster_map AS (
      SELECT NombreDeFantasia, Clusterizacion
      FROM `temple-brewery.Destileria.vw_ventas_con_cluster`
      WHERE Clusterizacion IS NOT NULL
      QUALIFY ROW_NUMBER() OVER (
        PARTITION BY NombreDeFantasia ORDER BY FechaPedido DESC
      ) = 1
    )
    SELECT
        FORMAT_DATE('%Y-%m-%d', ci.fecha_emision)                              AS f,
        COALESCE(cl.cluster, cm.Clusterizacion, 'Sin Cluster')                AS cl,
        COALESCE(NULLIF(TRIM(cl.nombre_fantasia),''), NULLIF(TRIM(cl.razon_social),''), 'Sin nombre') AS nd,
        ''                                                                      AS gc,
        COALESCE(ci.concepto, '')                                               AS pr,
        CASE
            WHEN UPPER(ci.concepto) LIKE '%473%'
              OR UPPER(ci.concepto) LIKE '%LATA%'  THEN 'LATA'
            WHEN UPPER(ci.concepto) LIKE '%BARRIL%' THEN 'BARRIL'
            ELSE ''
        END                                                                     AS en,
        COALESCE(ci.tipo_fc, '')                                                AS ti,
        (CASE WHEN ci.tipo_fc LIKE 'NC%' THEN -1 ELSE 1 END)
            * COALESCE(ci.cantidad,    0.0)                                     AS ce,
        (CASE WHEN ci.tipo_fc LIKE 'NC%' THEN -1 ELSE 1 END)
            * ROUND(COALESCE(ci.litros,     0.0), 3)                            AS li,
        (CASE WHEN ci.tipo_fc LIKE 'NC%' THEN -1 ELSE 1 END)
            * ROUND(COALESCE(ci.neto_linea, 0.0), 0)                            AS to_
    FROM `temple-bar-439715.Destileria_Contabilium.comprobantes_items` ci
    LEFT JOIN `temple-bar-439715.Destileria_Contabilium.clientes` cl
        ON ci.id_cliente = cl.id_cliente
    LEFT JOIN cluster_map cm
        ON TRIM(cl.nombre_fantasia) = cm.NombreDeFantasia
    WHERE ci.fecha_emision >= '2026-07-01'
      -- FCA/FCB/FCC/FCE/FCM = facturas reales; COT = venta entregada aún no facturada (a pedido del usuario);
      -- NCA/NCB/NCC/NCT = notas de crédito (se restan via el signo arriba)
      AND ci.tipo_fc IN ('FCA','FCB','FCC','FCE','FCM','COT','NCA','NCB','NCC','NCT')
      AND COALESCE(ci.cantidad, 0) > 0
      AND COALESCE(ci.tipo_item, 'P') != 'S'
    ORDER BY ci.fecha_emision
    """
    rows = list(client.query(query).result(timeout=60))
    result = []
    for r in rows:
        nd = r.nd
        cl = _CBL_CLUSTER_MAP.get(nd, r.cl)   # override cluster si el nombre está mapeado
        result.append(types.SimpleNamespace(
            f=r.f, cl=cl, nd=nd, gc=r.gc,
            pr=r.pr, en=r.en, ti=r.ti,
            ce=r.ce, li=r.li, to_=r.to_
        ))
    return result


# ---------------------------------------------------------------------------
# Validación de integridad Contabilium (evita publicar facturación incompleta)
# ---------------------------------------------------------------------------
#
# Incidente 2026-08-31: una corrida manual capturó BigQuery en un estado
# transitorio (probablemente el sync Contabilium→BQ recién había hecho el
# DELETE-antes-de-refetch de las COT, ver lesson_contabilium_sync_fixes) y
# publicó un tablero con 0 filas tipo COT — julio quedó en $37.2M en vez de
# $58.9M reales (-37%). Estos checks comparan lo recién obtenido de BQ contra
# el MÁXIMO HISTÓRICO conocido (no solo el último publicado — ver high-water-
# mark abajo) y abortan la publicación (sin subir nada) si detectan una
# regresión sospechosa, para nunca pisar un tablero correcto con uno peor.
# Si abortan, además dejan una bandera en GCS que el propio tablero en vivo
# muestra como banner (ver app.py) — no depende de que alguien mire un log.

_CBL_DROP_THRESHOLD = 0.10  # caída >10% en un mes vs. el máximo histórico = sospechoso
_CBL_HWM_BLOB       = "destileria_cbl_hwm.json"          # high-water-mark persistente
_CBL_ALERT_BLOB     = "destileria_alert.json"            # bandera de falla que lee app.py
_CBL_LAST_GOOD_BLOB = "destileria_dashboard_last_good.html"  # última publicación validada limpia


def is_contabilium_data_healthy(html_text, hwm):
    """Defensa en profundidad: valida un HTML de destileria_dashboard.html
    YA PUBLICADO en GCS contra el high-water-mark, sin importar qué proceso
    lo haya subido (el script correcto, una copia vieja, lo que sea). Usada
    por app.py en el momento de SERVIR la página — no solo por este script
    al momento de publicar — para que ninguna vía de publicación, conocida o
    no, pueda mostrarle a un usuario un número que ya sabemos que es peor
    que el máximo histórico. Devuelve (ok: bool, reasons: list[str])."""
    cur = _extract_contabilium_monthly(html_text) or {}
    reasons = []
    for mes, base in (hwm or {}).items():
        cur_b = cur.get(mes, {"total": 0.0, "cot": 0, "li": 0.0, "ce": 0.0})
        for campo, etiqueta in (("total", "Facturación $"), ("li", "Litros"), ("ce", "Envases")):
            b = base.get(campo, 0.0)
            if b > 0:
                drop = (b - cur_b.get(campo, 0.0)) / b
                if drop > _CBL_DROP_THRESHOLD:
                    reasons.append(
                        f"{etiqueta} Contabilium {mes} cayó {drop*100:.1f}% vs. el máximo histórico "
                        f"({b:,.0f} → {cur_b.get(campo, 0.0):,.0f})"
                    )
        if base.get("cot", 0) > 0 and cur_b.get("cot", 0) == 0:
            reasons.append(
                f"Contabilium {mes}: {base['cot']} comprobantes tipo COT vistos históricamente, "
                f"0 en lo publicado ahora — datos probablemente incompletos"
            )
    return (not reasons), reasons


def _extract_contabilium_monthly(html_text):
    """Parsea un HTML de destileria_dashboard.html ya generado y devuelve
    {mes: {'total': float, 'cot': int, 'items': int, 'li': float, 'ce': float}}
    para filas con fecha >= 2026-07-01 (período servido por Contabilium).
    None si no se pudo parsear (formato viejo, corrupto, etc.) — nunca lanza
    excepción. Usa json.raw_decode (no regex) para no truncar si algún campo
    de texto contuviera algo parecido a '];\\n'."""
    marker = "const ROWS = "
    i = html_text.find(marker)
    if i == -1:
        return None
    try:
        rows, _ = json.JSONDecoder().raw_decode(html_text, i + len(marker))
    except Exception:
        return None
    out = {}
    for r in rows:
        f = r.get("f") or ""
        if f < "2026-07-01":
            continue
        mes = f[:7]
        b = out.setdefault(mes, {"total": 0.0, "cot": 0, "items": 0, "li": 0.0, "ce": 0.0})
        b["total"] += r.get("to", 0.0)
        b["li"]    += r.get("li", 0.0)
        b["ce"]    += r.get("ce", 0.0)
        b["items"] += 1
        if r.get("ti") == "COT":
            b["cot"] += 1
    return out


def _monthly_from_raw(raw_cbl):
    """Mismo agregado que _extract_contabilium_monthly pero a partir de la
    lista de SimpleNamespace recién traída de BQ (fetch_rows_contabilium),
    no de un HTML ya generado."""
    cur = {}
    for r in raw_cbl:
        if not r.f:
            continue
        mes = r.f[:7]
        b = cur.setdefault(mes, {"total": 0.0, "cot": 0, "items": 0, "li": 0.0, "ce": 0.0})
        b["total"] += r.to_
        b["li"]    += r.li
        b["ce"]    += r.ce
        b["items"] += 1
        if r.ti == "COT":
            b["cot"] += 1
    return cur


def load_cbl_hwm(bucket_name):
    """High-water-mark: el máximo histórico ya validado para cada mes de
    Contabilium (persiste en GCS, sobrevive entre corridas). Comparar contra
    esto — y no solo contra el último tablero publicado — evita que una
    seguidilla de caídas chicas (cada una por debajo del umbral) vaya
    'reseteando' silenciosamente la base de comparación hasta perder de
    vista una caída grande acumulada."""
    if not bucket_name:
        return {}
    try:
        from google.cloud import storage
        client = storage.Client()
        blob = client.bucket(bucket_name).blob(_CBL_HWM_BLOB)
        if not blob.exists():
            return {}
        return json.loads(blob.download_as_text(encoding="utf-8"))
    except Exception as _e:
        print(f"WARN: no se pudo leer el high-water-mark de Contabilium — {_e}", file=sys.stderr)
        return {}


def save_cbl_hwm(bucket_name, hwm):
    if not bucket_name:
        return
    try:
        from google.cloud import storage
        client = storage.Client()
        blob = client.bucket(bucket_name).blob(_CBL_HWM_BLOB)
        blob.upload_from_string(
            json.dumps(hwm, ensure_ascii=False, indent=2).encode("utf-8"),
            content_type="application/json; charset=utf-8",
        )
        blob.cache_control = "no-cache, no-store, must-revalidate"
        blob.patch()
    except Exception as _e:
        print(f"WARN: no se pudo guardar el high-water-mark de Contabilium — {_e}", file=sys.stderr)


def write_alert_flag(bucket_name, reason):
    """Deja una bandera visible en GCS (destileria_alert.json) para que
    app.py la muestre como banner en el tablero EN VIVO. Una falla del
    pipeline local (Task Scheduler, sin nadie mirando dashboard_errors.log)
    así se ve directamente en la página que la gente realmente consulta,
    no solo en un archivo de log."""
    if not bucket_name:
        return
    try:
        from google.cloud import storage
        client = storage.Client()
        blob = client.bucket(bucket_name).blob(_CBL_ALERT_BLOB)
        payload = {"failed_at": datetime.now().strftime("%d/%m/%Y %H:%M"), "reason": reason}
        blob.upload_from_string(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            content_type="application/json; charset=utf-8",
        )
        blob.cache_control = "no-cache, no-store, must-revalidate"
        blob.patch()
    except Exception as _e:
        print(f"WARN: no se pudo escribir la bandera de alerta en GCS — {_e}", file=sys.stderr)


def clear_alert_flag(bucket_name):
    """Borra la bandera de alerta — se llama después de una publicación
    exitosa para que el banner desaparezca solo en cuanto se resuelve."""
    if not bucket_name:
        return
    try:
        from google.cloud import storage
        client = storage.Client()
        blob = client.bucket(bucket_name).blob(_CBL_ALERT_BLOB)
        if blob.exists():
            blob.delete()
    except Exception as _e:
        print(f"WARN: no se pudo borrar la bandera de alerta en GCS — {_e}", file=sys.stderr)


def validate_contabilium_integrity(raw_cbl, bucket_name):
    """Compara los totales mensuales de Contabilium (facturación, litros,
    envases, presencia de COT) recién obtenidos contra el máximo entre lo
    publicado AHORA MISMO en GCS y el high-water-mark histórico.

    Devuelve (errors, skip_note):
      - errors: lista de strings de error (vacía si todo OK). Cualquier
        error debe abortar la publicación en el llamador.
      - skip_note: None si la validación corrió normalmente (con o sin
        errores). String describiendo por qué NO se pudo validar del todo
        (ej: falla de lectura de GCS) — el llamador debe tratarlo como una
        alerta visible aunque no bloquee la publicación (fail-open pero
        NUNCA silencioso)."""
    errors = []
    skip_note = None
    if not bucket_name:
        return errors, "sin --gcs-bucket — validación de integridad salteada"

    prev = {}
    try:
        from google.cloud import storage
        client = storage.Client()
        blob = client.bucket(bucket_name).blob("destileria_dashboard.html")
        if blob.exists():
            prev_html = blob.download_as_text(encoding="utf-8")
            prev = _extract_contabilium_monthly(prev_html) or {}
    except Exception as _e:
        skip_note = f"no se pudo leer el tablero publicado para validar integridad ({_e})"
        print(f"WARN: {skip_note}", file=sys.stderr)

    hwm = load_cbl_hwm(bucket_name)

    baseline = {}
    for mes in set(prev) | set(hwm):
        p, h = prev.get(mes, {}), hwm.get(mes, {})
        baseline[mes] = {
            "total": max(p.get("total", 0.0), h.get("total", 0.0)),
            "cot":   max(p.get("cot", 0),     h.get("cot", 0)),
            "li":    max(p.get("li", 0.0),    h.get("li", 0.0)),
            "ce":    max(p.get("ce", 0.0),    h.get("ce", 0.0)),
        }
    if not baseline:
        return errors, skip_note

    cur = _monthly_from_raw(raw_cbl)

    for mes, base in sorted(baseline.items()):
        cur_b = cur.get(mes, {"total": 0.0, "cot": 0, "li": 0.0, "ce": 0.0})
        for campo, etiqueta in (("total", "Facturación $"), ("li", "Litros"), ("ce", "Envases")):
            if base[campo] > 0:
                drop = (base[campo] - cur_b[campo]) / base[campo]
                if drop > _CBL_DROP_THRESHOLD:
                    errors.append(
                        f"{etiqueta} Contabilium {mes} cayó {drop*100:.1f}% vs. el máximo histórico "
                        f"({base[campo]:,.0f} → {cur_b[campo]:,.0f})"
                    )
        if base["cot"] > 0 and cur_b["cot"] == 0:
            errors.append(
                f"Contabilium {mes}: {base['cot']} comprobantes tipo COT vistos históricamente, "
                f"0 en esta corrida — datos de BQ probablemente incompletos"
            )
    return errors, skip_note


def check_comprobantes_duplicates(client):
    """Detecta duplicados en comprobantes_items (bug histórico: streaming
    buffer + DELETE/reinsert de COT solapados duplicaba líneas de ítem y
    duplicaba la facturación, ver lesson_contabilium_sync_fixes Causa 3).
    Devuelve la cantidad de filas duplicadas (0 = OK)."""
    query = """
    SELECT COUNT(*) - COUNT(DISTINCT id_item) AS dup
    FROM `temple-bar-439715.Destileria_Contabilium.comprobantes_items`
    WHERE fecha_emision >= '2026-07-01'
    """
    row = list(client.query(query).result(timeout=60))[0]
    return int(row.dup or 0)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generar Destilería Dashboard")
    parser.add_argument("--output",     default=OUTPUT_DEFAULT,
                        help="Ruta del HTML generado")
    parser.add_argument("--gcs-bucket", default="",
                        help="Bucket GCS destino (opcional)")
    parser.add_argument("--force-publish", action="store_true",
                        help="Publica igual aunque los guardrails de integridad Contabilium "
                             "detecten una caída sospechosa. Usar solo tras confirmar manualmente "
                             "que el cambio es legítimo (ej: nota de crédito grande real).")
    args = parser.parse_args()

    ts = lambda: datetime.now().strftime("%H:%M:%S")

    if not args.gcs_bucket:
        print(f"[{ts()}] WARN: corriendo sin --gcs-bucket — no se va a publicar nada "
              f"y los guardrails de integridad Contabilium quedan salteados (nada contra qué comparar).")

    print(f"[{ts()}] Autenticando con Google (BQ + Sheets)...")
    creds, _ = google.auth.default(scopes=GCP_SCOPES)

    print(f"[{ts()}] Conectando a BigQuery (proyecto: {PROJECT})...")
    client = bigquery.Client(project=PROJECT, credentials=creds)

    print(f"[{ts()}] Consultando datos históricos...")
    raw = fetch_rows(client)
    print(f"[{ts()}] {len(raw):,} filas históricas")

    print(f"[{ts()}] Consultando Contabilium (desde 2026-07-01)...")
    raw_cbl = fetch_rows_contabilium(creds)
    print(f"[{ts()}] {len(raw_cbl):,} filas Contabilium")

    if raw_cbl:
        _sc_n = sum(1 for r in raw_cbl if r.cl == "Sin Cluster")
        print(f"[{ts()}] Info: {_sc_n}/{len(raw_cbl)} filas Contabilium sin cluster asignado "
              f"({_sc_n / len(raw_cbl) * 100:.0f}%) — no afecta la facturación total, "
              f"pero si es alto conviene revisar el panel admin de clusterización")

    # ── Guardrails de integridad Contabilium (ver incidente 2026-08-31) ─────
    print(f"[{ts()}] Verificando duplicados en comprobantes_items...")
    import google.cloud.bigquery as _bq
    _cbl_client = _bq.Client(project="temple-bar-439715", credentials=creds)
    _dup_count = check_comprobantes_duplicates(_cbl_client)
    if _dup_count > 0:
        _msg = (f"{_dup_count} filas duplicadas en Destileria_Contabilium.comprobantes_items "
                "(id_item repetido) — la facturación quedaría inflada.")
        if args.force_publish:
            print(f"[{ts()}] WARN --force-publish: se ignora el guardrail de duplicados — {_msg}", file=sys.stderr)
        else:
            write_alert_flag(args.gcs_bucket, _msg)
            raise RuntimeError(
                f"ERROR CRÍTICO: {_msg} Abortando sin publicar. Revisar sync Contabilium→BQ "
                "(streaming buffer / DELETE-reinsert solapados). Si es un falso positivo confirmado, "
                "volver a correr con --force-publish."
            )
    else:
        print(f"[{ts()}] Sin duplicados en comprobantes_items — OK")

    print(f"[{ts()}] Verificando integridad vs. máximo histórico...")
    _cbl_errors, _cbl_skip_note = validate_contabilium_integrity(raw_cbl, args.gcs_bucket)
    if _cbl_errors:
        for _e in _cbl_errors:
            print(f"[{ts()}] ERROR CRÍTICO: {_e}", file=sys.stderr)
        _detalle = " | ".join(_cbl_errors)
        if args.force_publish:
            print(f"[{ts()}] WARN --force-publish: se ignora el guardrail de integridad — {_detalle}", file=sys.stderr)
            write_alert_flag(args.gcs_bucket, f"Publicado con --force-publish pese a: {_detalle}")
        else:
            write_alert_flag(args.gcs_bucket, _detalle)
            raise RuntimeError(
                "Datos de Contabilium incompletos/inconsistentes respecto al máximo histórico — "
                "abortando para no sobrescribir con números peores. Revisar sync Contabilium→BQ "
                "(¿corrió completo antes de esta generación?) y volver a correr. "
                f"Detalle: {_detalle}. Si es un cambio legítimo confirmado, volver a correr con --force-publish."
            )
    elif _cbl_skip_note:
        print(f"[{ts()}] WARN: validación de integridad incompleta — {_cbl_skip_note}")
        write_alert_flag(args.gcs_bucket, f"Guardrail no pudo validar del todo esta corrida: {_cbl_skip_note}")
    else:
        print(f"[{ts()}] Integridad Contabilium OK")

    raw = raw + raw_cbl

    # Alerta de datos desactualizados: si la última fecha sincronizada de
    # Contabilium queda muy atrás de hoy, el sync Contabilium→BQ probablemente
    # viene fallando en silencio (ver incidente 2026-08-18: 4 días sin correr).
    stale_banner = ""
    _cbl_dates = [r.f for r in raw_cbl if r.f]
    if _cbl_dates:
        _last_cbl_date = max(_cbl_dates)
        _days_stale = (datetime.now().date() - datetime.strptime(_last_cbl_date, "%Y-%m-%d").date()).days
        if _days_stale > 2:
            stale_banner = (
                '<div style="background:#3d1f1f;border:1px solid #c8102e;border-radius:8px;'
                'padding:12px 18px;margin:0 auto 14px;max-width:1400px;color:#fca5a5;'
                'font-size:13px;font-weight:600">'
                f'⚠ Datos de Contabilium desactualizados — última sincronización: {_last_cbl_date} '
                f'({_days_stale} días de atraso). Revisar el sync Contabilium→BQ / tareas programadas.'
                '</div>'
            )
            print(f"[{ts()}] ALERTA: Contabilium desactualizado ({_days_stale} días, última fecha {_last_cbl_date})")

    print(f"[{ts()}] Total combinado: {len(raw):,} filas")
    if len(raw) < 100:
        raise RuntimeError(f"BQ devolvió solo {len(raw)} filas — posible error de datos, abortando")

    # Construir array compacto para el frontend
    data = []
    for r in raw:
        data.append({
            "f":  r.f,
            "cl": r.cl,
            "nd": r.nd,
            "fa": classify_familia(r.pr, r.en),
            "ti": r.ti,
            "ce": int(r.ce),
            "li": float(r.li),
            "to": float(r.to_),
        })

    now_str   = datetime.now().strftime("%d/%m/%Y %H:%M")
    rows_json = json.dumps(data, ensure_ascii=False, separators=(",", ":"))

    # Leer plantilla e inyectar
    if not os.path.exists(TEMPLATE):
        print(f"ERROR: No se encontro la plantilla: {TEMPLATE}", file=sys.stderr)
        sys.exit(1)

    with open(TEMPLATE, encoding="utf-8") as fh:
        template = fh.read()

    # ── Validación estructural crítica ───────────────────────────────────────
    # Estas secciones NO deben modificarse. Si el template pierde alguna (ej:
    # Google Drive sync baja una versión vieja), el pipeline aborta antes de
    # publicar un dashboard roto.
    _CRITICAL_CHECKS = [
        (
            'data-tab="feriado-semanas"',
            "Tab Semanas de Feriado (botón de navegación)",
        ),
        (
            'id="view-feriado-semanas"',
            "Vista Feriado Semanas (#view-feriado-semanas)",
        ),
        (
            'id="bosque-obj-panel"',
            "Panel de objetivos Bosque (#bosque-obj-panel)",
        ),
        (
            'id="bosque-retention-panel"',
            "Panel retención Bosque (#bosque-retention-panel)",
        ),
        (
            'data-tab="bosque-sellinout"',
            "Tab Sell In/Out Bosque (botón de navegación)",
        ),
        (
            'id="view-bosque-sellinout"',
            "Vista Sell In/Out Bosque (#view-bosque-sellinout)",
        ),
        (
            '__SELLINOUT_LOCAL_JSON__',
            "Placeholder sell-in/out por local (__SELLINOUT_LOCAL_JSON__)",
        ),
        (
            '__SELLINOUT_LOCAL_WK_JSON__',
            "Placeholder sell-in/out local × semana (__SELLINOUT_LOCAL_WK_JSON__)",
        ),
        (
            '__SELLINOUT_LOCAL_MO_JSON__',
            "Placeholder sell-in/out local × mes (__SELLINOUT_LOCAL_MO_JSON__)",
        ),
        (
            '__SELLINOUT_PAT_JSON__',
            "Placeholder sell-in/out Patagonia semanal (__SELLINOUT_PAT_JSON__)",
        ),
        (
            '__SELLINOUT_PAT_LOCAL_WK_JSON__',
            "Placeholder sell-in/out Patagonia local × semana (__SELLINOUT_PAT_LOCAL_WK_JSON__)",
        ),
        (
            '__SELLINOUT_PAT_LOCAL_MO_JSON__',
            "Placeholder sell-in/out Patagonia local × mes (__SELLINOUT_PAT_LOCAL_MO_JSON__)",
        ),
        (
            '__SELLINOUT_FER_JSON__',
            "Placeholder sell-in/out Feriado semanal (__SELLINOUT_FER_JSON__)",
        ),
        (
            '__SELLINOUT_FER_LOCAL_WK_JSON__',
            "Placeholder sell-in/out Feriado local × semana (__SELLINOUT_FER_LOCAL_WK_JSON__)",
        ),
        (
            '__SELLINOUT_FER_LOCAL_MO_JSON__',
            "Placeholder sell-in/out Feriado local × mes (__SELLINOUT_FER_LOCAL_MO_JSON__)",
        ),
        (
            '__SELLINOUT_FER_PAT_JSON__',
            "Placeholder sell-in/out Feriado Patagonia semanal (__SELLINOUT_FER_PAT_JSON__)",
        ),
        (
            '__SELLINOUT_FER_PAT_LOCAL_WK_JSON__',
            "Placeholder sell-in/out Feriado Patagonia local × semana (__SELLINOUT_FER_PAT_LOCAL_WK_JSON__)",
        ),
        (
            '__SELLINOUT_FER_PAT_LOCAL_MO_JSON__',
            "Placeholder sell-in/out Feriado Patagonia local × mes (__SELLINOUT_FER_PAT_LOCAL_MO_JSON__)",
        ),
        (
            '__PERMISSIONS_INJECT__',
            "Placeholder inyección de permisos y link Admin (__PERMISSIONS_INJECT__)",
        ),
    ]
    _tpl_errors = []
    for pattern, label in _CRITICAL_CHECKS:
        if pattern not in template:
            _tpl_errors.append(f"  ✗ FALTA: {label}  →  buscar: {pattern!r}")
    # Verificar orden: objetivos ANTES que retención en Bosque General
    _idx_obj = template.find('id="bosque-obj-panel"')
    _idx_ret = template.find('id="bosque-retention-panel"')
    if _idx_obj != -1 and _idx_ret != -1 and _idx_obj > _idx_ret:
        _tpl_errors.append(
            "  ✗ ORDEN INCORRECTO: bosque-obj-panel debe aparecer ANTES de bosque-retention-panel"
        )
    if _tpl_errors:
        print(
            f"\n{'!' * 60}\n"
            f"ERROR CRÍTICO: El template tiene secciones faltantes o en orden incorrecto.\n"
            f"Posible causa: Google Drive sincronizó una versión vieja del template.\n"
            + "\n".join(_tpl_errors)
            + f"\n{'!' * 60}\n",
            file=sys.stderr,
        )
        sys.exit(2)
    print(f"[{ts()}] Template OK — validación estructural superada ({len(_CRITICAL_CHECKS)} checks)")
    # ── Fin validación ───────────────────────────────────────────────────────

    print(f"[{ts()}] Leyendo objetivos (Firestore → Drive → GCS cache → JSON local → vacío)...")

    # Cargar caché anterior para comparar después del fetch
    _prev_obj = None
    if os.path.exists(OBJ_JSON_FILE):
        try:
            with open(OBJ_JSON_FILE, encoding="utf-8") as fh:
                _cached = json.load(fh)
            _prev_obj = {k: v for k, v in _cached.items() if k != "_meta"}
        except Exception:
            pass

    def _stale_days(meta):
        try:
            return (datetime.now() - datetime.fromisoformat(meta["fetched_at"])).days
        except Exception:
            return "?"

    def _apply_cached(cached_dict, source_label):
        _meta = cached_dict.pop("_meta", {})
        print(
            f"[{ts()}] Objetivos desde {source_label} — FALLBACK "
            f"(antigüedad: {_stale_days(_meta)} día(s), "
            f"última sync: {_meta.get('fetched_at', 'desconocida')})",
            file=sys.stderr,
        )
        return cached_dict

    obj = None
    _obj_source = "none"

    # ── 0. Intentar desde Firestore (con reintento si la lectura viene incompleta) ──
    try:
        from google.cloud import firestore as _firestore
        _fs_client = _firestore.Client(project="temple-bar-439715")
        _fs_obj = load_objectives_from_firestore(_fs_client)
        _fs_errs, _ = validate_objectives(_fs_obj) if _fs_obj else ([], [])
        if _fs_obj and _fs_errs:
            print(f"[{ts()}] WARN: lectura de Firestore incompleta ({len(_fs_errs)} error(es)) — reintentando...", file=sys.stderr)
            time.sleep(3)
            _fs_obj = load_objectives_from_firestore(_fs_client)
            _fs_errs, _ = validate_objectives(_fs_obj) if _fs_obj else ([], [])
        if _fs_obj and not _fs_errs:
            obj = _fs_obj
            _obj_source = "firestore"
            print(f"[{ts()}] Objetivos OK (desde Firestore)")
        elif _fs_obj:
            print(f"[{ts()}] WARN: Firestore sigue incompleto tras reintento — se descarta, se prueba el siguiente fallback", file=sys.stderr)
    except Exception as _fs_init_err:
        print(f"WARN: Firestore init falló: {_fs_init_err}", file=sys.stderr)

    # ── 1. Intentar desde Drive ──────────────────────────────────────────────
    if obj is None:
        try:
            obj = fetch_objectives(creds)
            _obj_source = "drive"
            print(f"[{ts()}] Objetivos OK (desde Drive)")
            _to_save = {**obj, "_meta": {
                "source": "drive",
                "fetched_at": datetime.now().isoformat(timespec="seconds"),
                "file_id": SHEET_OBJ_ID,
            }}
            # Guardar en local
            with open(OBJ_JSON_FILE, "w", encoding="utf-8") as fh:
                json.dump(_to_save, fh, ensure_ascii=False, indent=2)
            print(f"[{ts()}] JSON local de objetivos actualizado")
            # Guardar en GCS (persiste entre runs de Cloud Run)
            if args.gcs_bucket:
                try:
                    save_objectives_to_gcs(_to_save, args.gcs_bucket)
                except Exception as _gcs_save_err:
                    print(f"WARN: No se pudo guardar objetivos en GCS: {_gcs_save_err}", file=sys.stderr)
        except Exception as exc:
            import traceback
            print(f"WARN: Drive falló: {exc}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)

    # ── 2. Fallback: GCS cache (sobrevive reinicios de Cloud Run) ───────────
    if obj is None and args.gcs_bucket:
        try:
            _cached = load_objectives_from_gcs(args.gcs_bucket)
            obj = _apply_cached(_cached, f"gs://{args.gcs_bucket}/{OBJ_GCS_BLOB}")
            _obj_source = "gcs_cache"
        except Exception as _gcs_err:
            print(f"WARN: GCS cache falló: {_gcs_err}", file=sys.stderr)

    # ── 3. Fallback: archivo local ───────────────────────────────────────────
    if obj is None and os.path.exists(OBJ_JSON_FILE):
        try:
            with open(OBJ_JSON_FILE, encoding="utf-8") as fh:
                _cached = json.load(fh)
            obj = _apply_cached(_cached, os.path.basename(OBJ_JSON_FILE))
            _obj_source = "local_cache"
        except Exception as _local_err:
            print(f"WARN: Caché local falló: {_local_err}", file=sys.stderr)

    # ── 4. Sin objetivos — dashboard se genera con aviso visible ────────────
    _obj_missing = obj is None
    if _obj_missing:
        obj = {b: {"product": {}, "cluster": {}} for b in ("bosque", "feriado", "cerveza")}
        _obj_source = "empty"
        print("WARN: Sin objetivos disponibles — dashboard sin sección de objetivos", file=sys.stderr)

    # Validar estructura y detectar anomalías (solo log, nunca abortar)
    _errs, _warns = validate_objectives(obj, _prev_obj)
    if _errs or _warns:
        _hdr = (
            f"\n{'!' * 60}\n"
            f"⚠  ALERTA OBJETIVOS DESTILERÍA — fuente: {_obj_source}"
            + (f" | {len(_errs)} ERROR(ES)" if _errs else "")
            + (f" | {len(_warns)} cambio(s) grandes" if _warns else "")
            + f"\n"
        )
        print(_hdr, file=sys.stderr)
        for e in _errs:
            print(f"  ❌ ERROR  : {e}", file=sys.stderr)
        for w in _warns:
            print(f"  ⚠  CAMBIO: {w}", file=sys.stderr)
        print(f"{'!' * 60}\n", file=sys.stderr)
        # Si Drive devolvió objetivos inválidos, intentar GCS cache como recuperación
        if _errs and _obj_source == "drive" and args.gcs_bucket:
            try:
                _cached = load_objectives_from_gcs(args.gcs_bucket)
                _cached_errs, _ = validate_objectives(_cached)
                if not _cached_errs:
                    obj = _apply_cached(_cached, f"gs://{args.gcs_bucket}/{OBJ_GCS_BLOB} [recuperación]")
                    _obj_source = "gcs_cache"
                    _obj_missing = False
                    print(f"[{ts()}] Recuperado: usando GCS cache en lugar de objetivos inválidos de Drive")
            except Exception:
                pass

    obj_json  = json.dumps(obj, ensure_ascii=False, separators=(",", ":"))

    html = template.replace("__ROWS_JSON__",   rows_json)
    html = html.replace("__OBJ_JSON__",        obj_json)
    html = html.replace("__UPDATED_AT__",      now_str)
    html = html.replace("__RECORD_COUNT__",    f"{len(data):,}")
    html = html.replace("__STALE_BANNER__",    stale_banner)

    # ── Sell-in / Sell-out Bosque Nativo ────────────────────────────────────
    try:
        _sio = fetch_sellinout_weekly(creds, weeks=12)
        html = html.replace("__SELLINOUT_JSON__", json.dumps(_sio, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT inyectado ({len(_sio)} semanas)")
    except Exception as _sio_err:
        print(f"WARN: SELLINOUT falló — {_sio_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_JSON__", "[]")

    # ── Sell-in / Sell-out por local ────────────────────────────────────────
    try:
        _sio_local = fetch_sellinout_by_local(creds, weeks=12)
        html = html.replace("__SELLINOUT_LOCAL_JSON__", json.dumps(_sio_local, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_LOCAL inyectado ({len(_sio_local)} locales)")
    except Exception as _sio_local_err:
        print(f"WARN: SELLINOUT_LOCAL falló — {_sio_local_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_LOCAL_JSON__", "[]")

    # ── Sell-in / Sell-out por local × semana ───────────────────────────────
    try:
        _sio_lw = fetch_sellinout_local_weekly(creds, weeks=12)
        html = html.replace("__SELLINOUT_LOCAL_WK_JSON__", json.dumps(_sio_lw, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_LOCAL_WK inyectado ({len(_sio_lw)} filas)")
    except Exception as _sio_lw_err:
        print(f"WARN: SELLINOUT_LOCAL_WK falló — {_sio_lw_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_LOCAL_WK_JSON__", "[]")

    # ── Sell-in / Sell-out por local × mes calendario (Mes Actual/Anterior) ──
    try:
        _sio_lm = fetch_sellinout_local_monthly(
            creds, r'GIN BOSQUE.*(500|750)', "curated_gin", "Gin_Total",
            _SI_ALIAS, _SO_ALIAS, months=3,
        )
        html = html.replace("__SELLINOUT_LOCAL_MO_JSON__", json.dumps(_sio_lm, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_LOCAL_MO inyectado ({len(_sio_lm)} filas)")
    except Exception as _sio_lm_err:
        print(f"WARN: SELLINOUT_LOCAL_MO falló — {_sio_lm_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_LOCAL_MO_JSON__", "[]")

    try:
        _sio_pat = fetch_sellinout_pat_weekly(creds, weeks=12)
        html = html.replace("__SELLINOUT_PAT_JSON__", json.dumps(_sio_pat, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_PAT inyectado ({len(_sio_pat)} semanas)")
    except Exception as _sio_pat_err:
        print(f"WARN: SELLINOUT_PAT falló — {_sio_pat_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_PAT_JSON__", "[]")

    try:
        _sio_pat_lw = fetch_sellinout_pat_local_weekly(creds, weeks=12)
        html = html.replace("__SELLINOUT_PAT_LOCAL_WK_JSON__", json.dumps(_sio_pat_lw, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_PAT_LOCAL_WK inyectado ({len(_sio_pat_lw)} filas)")
    except Exception as _sio_pat_lw_err:
        print(f"WARN: SELLINOUT_PAT_LOCAL_WK falló — {_sio_pat_lw_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_PAT_LOCAL_WK_JSON__", "[]")

    try:
        _sio_pat_lm = fetch_sellinout_pat_local_monthly(creds, months=3)
        html = html.replace("__SELLINOUT_PAT_LOCAL_MO_JSON__", json.dumps(_sio_pat_lm, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_PAT_LOCAL_MO inyectado ({len(_sio_pat_lm)} filas)")
    except Exception as _sio_pat_lm_err:
        print(f"WARN: SELLINOUT_PAT_LOCAL_MO falló — {_sio_pat_lm_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_PAT_LOCAL_MO_JSON__", "[]")

    # ── Sell-in / Sell-out Feriado ───────────────────────────────────────────
    try:
        _sio_fer = fetch_sellinout_fer_weekly(creds, weeks=12)
        html = html.replace("__SELLINOUT_FER_JSON__", json.dumps(_sio_fer, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_FER inyectado ({len(_sio_fer)} semanas)")
    except Exception as _sio_fer_err:
        print(f"WARN: SELLINOUT_FER falló — {_sio_fer_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_FER_JSON__", "[]")

    try:
        _sio_fer_lw = fetch_sellinout_fer_local_weekly(creds, weeks=12)
        html = html.replace("__SELLINOUT_FER_LOCAL_WK_JSON__", json.dumps(_sio_fer_lw, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_FER_LOCAL_WK inyectado ({len(_sio_fer_lw)} filas)")
    except Exception as _sio_fer_lw_err:
        print(f"WARN: SELLINOUT_FER_LOCAL_WK falló — {_sio_fer_lw_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_FER_LOCAL_WK_JSON__", "[]")

    # ── Sell-in / Sell-out Feriado por local × mes calendario ────────────────
    try:
        _sio_fer_lm = fetch_sellinout_local_monthly(
            creds, r'FERIADO', "curated_feriado", "Feriado_Total",
            _FER_SI_ALIAS, _FER_SO_ALIAS, months=3,
        )
        html = html.replace("__SELLINOUT_FER_LOCAL_MO_JSON__", json.dumps(_sio_fer_lm, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_FER_LOCAL_MO inyectado ({len(_sio_fer_lm)} filas)")
    except Exception as _sio_fer_lm_err:
        print(f"WARN: SELLINOUT_FER_LOCAL_MO falló — {_sio_fer_lm_err}", file=sys.stderr)
        html = html.replace("__SELLINOUT_FER_LOCAL_MO_JSON__", "[]")

    try:
        _sio_fer_pat = fetch_sellinout_fer_pat_weekly(creds, weeks=12)
        html = html.replace("__SELLINOUT_FER_PAT_JSON__", json.dumps(_sio_fer_pat, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_FER_PAT inyectado ({len(_sio_fer_pat)} semanas)")
    except Exception as _e:
        print(f"WARN: SELLINOUT_FER_PAT falló — {_e}", file=sys.stderr)
        html = html.replace("__SELLINOUT_FER_PAT_JSON__", "[]")

    try:
        _sio_fer_pat_lw = fetch_sellinout_fer_pat_local_weekly(creds, weeks=12)
        html = html.replace("__SELLINOUT_FER_PAT_LOCAL_WK_JSON__", json.dumps(_sio_fer_pat_lw, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_FER_PAT_LOCAL_WK inyectado ({len(_sio_fer_pat_lw)} filas)")
    except Exception as _e:
        print(f"WARN: SELLINOUT_FER_PAT_LOCAL_WK falló — {_e}", file=sys.stderr)
        html = html.replace("__SELLINOUT_FER_PAT_LOCAL_WK_JSON__", "[]")

    try:
        _sio_fer_pat_lm = fetch_sellinout_fer_pat_local_monthly(creds, months=3)
        html = html.replace("__SELLINOUT_FER_PAT_LOCAL_MO_JSON__", json.dumps(_sio_fer_pat_lm, ensure_ascii=False, separators=(",", ":")))
        print(f"[{ts()}] ✓ SELLINOUT_FER_PAT_LOCAL_MO inyectado ({len(_sio_fer_pat_lm)} filas)")
    except Exception as _e:
        print(f"WARN: SELLINOUT_FER_PAT_LOCAL_MO falló — {_e}", file=sys.stderr)
        html = html.replace("__SELLINOUT_FER_PAT_LOCAL_MO_JSON__", "[]")

    # Garantía: __PERMISSIONS_INJECT__ siempre presente aunque Drive haya
    # sincronizado una versión vieja del template que no lo tenga.
    if "__PERMISSIONS_INJECT__" not in html:
        html = html.replace("<body>", "<body>\n__PERMISSIONS_INJECT__", 1)
        print(f"[{ts()}] WARN: __PERMISSIONS_INJECT__ no estaba en el template — insertado automáticamente")

    # Banner visible cuando los objetivos no están disponibles
    if _obj_missing:
        _banner = (
            '<div style="background:#b91c1c;color:#fff;text-align:center;padding:10px 16px;'
            'font-family:sans-serif;font-size:14px;font-weight:600;position:sticky;top:0;z-index:9999">'
            '⚠ Objetivos no disponibles — Drive no accesible al momento de generar este dashboard. '
            f'Generado: {now_str}'
            '</div>'
        )
        html = html.replace("<body", _banner + "<body", 1)

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as fh:
        fh.write(html)

    kb = len(html.encode("utf-8")) // 1024
    print(f"[{ts()}] Generado: {args.output} ({kb} KB · {len(data):,} registros)")

    if args.gcs_bucket:
        upload_to_gcs(args.output, args.gcs_bucket, html_content=html)
        try:
            _hwm = load_cbl_hwm(args.gcs_bucket)
            for _mes, _b in _monthly_from_raw(raw_cbl).items():
                _h = _hwm.get(_mes, {})
                _hwm[_mes] = {
                    "total": max(_b["total"], _h.get("total", 0.0)),
                    "cot":   max(_b["cot"],   _h.get("cot", 0)),
                    "li":    max(_b["li"],    _h.get("li", 0.0)),
                    "ce":    max(_b["ce"],    _h.get("ce", 0.0)),
                }
            save_cbl_hwm(args.gcs_bucket, _hwm)
            if not _cbl_skip_note:
                clear_alert_flag(args.gcs_bucket)
                # Copia de respaldo para que app.py pueda mostrarla si una
                # publicación futura (por esta vía o cualquier otra) resulta
                # incompleta — ver is_contabilium_data_healthy.
                from google.cloud import storage as _storage
                _storage.Client().bucket(args.gcs_bucket).blob(_CBL_LAST_GOOD_BLOB) \
                    .upload_from_string(html.encode("utf-8"), content_type="text/html; charset=utf-8")
                print(f"[{ts()}] ✓ Copia de respaldo actualizada ({_CBL_LAST_GOOD_BLOB})")
        except Exception as _e:
            print(f"WARN: no se pudo actualizar el high-water-mark / copia de respaldo de Contabilium — {_e}", file=sys.stderr)
    else:
        print()
        print("Deploy:")
        print('  gsutil -h "Cache-Control:no-cache, no-store, must-revalidate" '
              'cp destileria_dashboard.html '
              'gs://temple-bar-dashboard-cache/destileria_dashboard.html')


# ── Sell-in (compra) desde Contabilium — reemplaza Ventas_Maestro_Con_Cluster_Final
# desde el 2026-07-01 (tabla congelada, sin filas nuevas desde esa fecha).
# Nombres de cliente en Contabilium != NombreDeFantasia de Ventas_Maestro, de ahí este alias aparte.
_CBL_SI_ALIAS: dict[str, str | None] = {
    "TEMPLE PUERTO MADERO":   "Puerto Madero",
    "TEMPLE HOLLYWOOD NUEVO": "Hollywood",
    "TEMPLE MASCHWITZ":       "Maschwitz",
    "TEMPLE PALERMO NUEVO":   "Casa Temple",
    "TEMPLE PASEO LA PLAZA":  "Club Temple",
    "TEMPLE RECOLETA SRL":    "Recoleta",
    "TEMPLE RIO GALLEGOS":    "Rio Gallegos",
    "TEMPLE ROSARIO":         "Rosario 2",
    "TEMPLE SALTA":           "Salta",
    "TEMPLE SOHO":            "Soho",
    "TEMPLE CAMINITO":        "Caminito",
    # nombre_fantasia mal cargado en Contabilium (dice "Cordoba" pero la razón
    # social MOLBRA S.A.S. es en realidad el local Güemes — confirmado 2026-08-11).
    "TEMPLE CORDOBA (USAR)":  "Güemes",
    "BONALAR S.R.L":                      "Santiago del Estero",
    "AREA SUR S.R.L. EN FORMACION":        "Comodoro Rivadavia",
    "QUIJOTE TUCUMAN  S. R. L.":           "Tucumán 3",
    # Los Templos Caballito S.R.L. es la razón social de facturación compartida
    # entre Barrio Chino y Monroe — no se puede separar la compra por local.
    "LOS TEMPLOS CABALLITO S.R.L.": "Barrio Chino + Monroe",
}

_CBL_SELLIN_CUTOFF = "2026-07-01"


def _fetch_cbl_sellin_by_week_local(creds, product_regex, alias_map=None,
                                     cluster_name="Cadena Grupo Temple", since=_CBL_SELLIN_CUTOFF):
    """Compra (sell-in) desde Contabilium, agrupada por semana y local.

    `alias_map`/`cluster_name` permiten reusar la función para otras cadenas
    (ej. Cadena Patagonia con `_CBL_PAT_SI_ALIAS`) — por defecto Grupo Temple.
    Devuelve {(local_canónico, 'YYYY-MM-DD' lunes de semana): litros}.
    """
    alias_map = _CBL_SI_ALIAS if alias_map is None else alias_map
    from google.cloud import bigquery as _bq
    client = _bq.Client(project="temple-bar-439715", credentials=creds)
    query = f"""
    WITH cluster_map AS (
      SELECT NombreDeFantasia, Clusterizacion
      FROM `temple-brewery.Destileria.vw_ventas_con_cluster`
      WHERE Clusterizacion IS NOT NULL
      QUALIFY ROW_NUMBER() OVER (
        PARTITION BY NombreDeFantasia ORDER BY FechaPedido DESC
      ) = 1
    )
    SELECT
      DATE_TRUNC(ci.fecha_emision, WEEK(MONDAY)) AS semana,
      COALESCE(NULLIF(TRIM(cl.nombre_fantasia),''), NULLIF(TRIM(cl.razon_social),''), 'Sin nombre') AS nd,
      COALESCE(cl.cluster, cm.Clusterizacion, 'Sin Cluster') AS cluster_raw,
      (CASE WHEN ci.tipo_fc LIKE 'NC%' THEN -1 ELSE 1 END)
          * ROUND(COALESCE(ci.litros, 0.0), 3)   AS litros
    FROM `temple-bar-439715.Destileria_Contabilium.comprobantes_items` ci
    LEFT JOIN `temple-bar-439715.Destileria_Contabilium.clientes` cl
        ON ci.id_cliente = cl.id_cliente
    LEFT JOIN cluster_map cm
        ON TRIM(cl.nombre_fantasia) = cm.NombreDeFantasia
    WHERE ci.fecha_emision >= '{since}'
      AND ci.tipo_fc IN ('FCA','FCB','FCC','FCE','FCM','COT','NCA','NCB','NCC','NCT')
      AND COALESCE(ci.cantidad, 0) > 0
      AND COALESCE(ci.tipo_item, 'P') != 'S'
      AND REGEXP_CONTAINS(UPPER(TRIM(ci.concepto)), r'{product_regex}')
    """
    result: dict[tuple, float] = {}
    for r in client.query(query).result(timeout=60):
        # Nombres curados en alias_map son locales conocidos — confiar en el alias
        # directo, sin pasar por Clusterizacion (que puede venir "Sin Cluster" para
        # nombres que _CBL_CLUSTER_MAP todavía no cubre).
        if r.nd in alias_map:
            canon = alias_map[r.nd]
            if canon is None:
                continue
        else:
            cluster = _CBL_CLUSTER_MAP.get(r.nd, r.cluster_raw)
            if cluster != cluster_name:
                continue
            canon = r.nd.title()
        key = (canon, str(r.semana))
        result[key] = result.get(key, 0.0) + float(r.litros)
    return result


def _fetch_cbl_sellin_by_month_local(creds, product_regex, alias_map=None,
                                      cluster_name="Cadena Grupo Temple", since=_CBL_SELLIN_CUTOFF):
    """Igual que _fetch_cbl_sellin_by_week_local pero agrupado por mes calendario
    ('YYYY-MM'). Usado para Mes Actual/Mes Anterior — evita el sesgo de los buckets
    semanales que arrastran días del mes anterior cuando el 1° no cae lunes."""
    alias_map = _CBL_SI_ALIAS if alias_map is None else alias_map
    from google.cloud import bigquery as _bq
    client = _bq.Client(project="temple-bar-439715", credentials=creds)
    query = f"""
    WITH cluster_map AS (
      SELECT NombreDeFantasia, Clusterizacion
      FROM `temple-brewery.Destileria.vw_ventas_con_cluster`
      WHERE Clusterizacion IS NOT NULL
      QUALIFY ROW_NUMBER() OVER (
        PARTITION BY NombreDeFantasia ORDER BY FechaPedido DESC
      ) = 1
    )
    SELECT
      FORMAT_DATE('%Y-%m', ci.fecha_emision) AS mes,
      COALESCE(NULLIF(TRIM(cl.nombre_fantasia),''), NULLIF(TRIM(cl.razon_social),''), 'Sin nombre') AS nd,
      COALESCE(cl.cluster, cm.Clusterizacion, 'Sin Cluster') AS cluster_raw,
      (CASE WHEN ci.tipo_fc LIKE 'NC%' THEN -1 ELSE 1 END)
          * ROUND(COALESCE(ci.litros, 0.0), 3)   AS litros
    FROM `temple-bar-439715.Destileria_Contabilium.comprobantes_items` ci
    LEFT JOIN `temple-bar-439715.Destileria_Contabilium.clientes` cl
        ON ci.id_cliente = cl.id_cliente
    LEFT JOIN cluster_map cm
        ON TRIM(cl.nombre_fantasia) = cm.NombreDeFantasia
    WHERE ci.fecha_emision >= '{since}'
      AND ci.tipo_fc IN ('FCA','FCB','FCC','FCE','FCM','COT','NCA','NCB','NCC','NCT')
      AND COALESCE(ci.cantidad, 0) > 0
      AND COALESCE(ci.tipo_item, 'P') != 'S'
      AND REGEXP_CONTAINS(UPPER(TRIM(ci.concepto)), r'{product_regex}')
    """
    result: dict[tuple, float] = {}
    for r in client.query(query).result(timeout=60):
        if r.nd in alias_map:
            canon = alias_map[r.nd]
            if canon is None:
                continue
        else:
            cluster = _CBL_CLUSTER_MAP.get(r.nd, r.cluster_raw)
            if cluster != cluster_name:
                continue
            canon = r.nd.title()
        key = (canon, r.mes)
        result[key] = result.get(key, 0.0) + float(r.litros)
    return result


def fetch_sellinout_local_monthly(creds, product_regex, so_table, so_field,
                                   si_alias, so_alias, months=3):
    """Compra vs venta por local y MES CALENDARIO (Cadena Grupo Temple) — usado
    exclusivamente por los filtros Mes Actual / Mes Anterior en las vistas 'Por
    Local' y 'Sem × Local' para evitar el sesgo de los buckets semanales.
    """
    from google.cloud import bigquery as _bq
    bq_dest   = _bq.Client(project="temple-brewery",    credentials=creds)
    bq_temple = _bq.Client(project="temple-bar-439715", credentials=creds)

    q_si = f"""
    SELECT
      FORMAT_DATE('%Y-%m', FechaPedido) AS mes,
      TRIM(NombreDeFantasia)            AS local,
      ROUND(SUM(COALESCE(Litros, 0)), 2) AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena grupo temple'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'{product_regex}')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {months} MONTH)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY mes, local
    """
    q_so = f"""
    SELECT
      FORMAT_DATE('%Y-%m', Fecha) AS mes,
      TRIM(Establecimiento)       AS local,
      ROUND(SUM({so_field}), 2)   AS litros
    FROM `temple-bar-439715.curated_database.{so_table}`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {months} MONTH)
    GROUP BY mes, local
    """

    si_map: dict[tuple, float] = {}
    for r in bq_dest.query(q_si).result(timeout=60):
        canon = si_alias.get(r.local, r.local)
        if canon is None:
            continue
        key = (canon, r.mes)
        si_map[key] = si_map.get(key, 0.0) + float(r.litros)

    for (local, mes), litros in _fetch_cbl_sellin_by_month_local(creds, product_regex).items():
        key = (local, mes)
        si_map[key] = si_map.get(key, 0.0) + litros

    so_map: dict[tuple, float] = {}
    for r in bq_temple.query(q_so).result(timeout=60):
        raw = r.local or ""
        canon = so_alias.get(raw.upper(), raw.title())
        if canon is None:
            continue
        key = (canon, r.mes)
        so_map[key] = so_map.get(key, 0.0) + float(r.litros)

    result = []
    for (local, mes) in sorted(set(si_map) | set(so_map)):
        si = round(si_map.get((local, mes), 0.0), 2)
        so = round(so_map.get((local, mes), 0.0), 2)
        if si == 0 and so == 0:
            continue
        result.append({"local": local, "mo": mes, "si": si, "so": so})
    return result


def fetch_sellinout_weekly(creds, weeks=12):
    """
    Retorna lista de semanas (más reciente primero) con sell-in y sell-out
    de Bosque Nativo para el cluster Cadena Grupo Temple.

    Sell-in : Ventas_Maestro_Con_Cluster_Final — cl='CADENA GRUPO TEMPLE',
              fa='bosque_nativo' (cubre 500ml y 750ml), campo li (litros).
    Sell-out: temple-bar-439715.curated_database.curated_gin — todos los
              registros, campo Gin_Total (litros por serve ya calculado).

    Alarma: sell-out[semana N] > sell-in[semana N-1]
    """
    from google.cloud import bigquery as _bq
    bq_dest   = _bq.Client(project="temple-brewery",    credentials=creds)
    bq_temple = _bq.Client(project="temple-bar-439715", credentials=creds)

    q_si = f"""
    SELECT
      DATE_TRUNC(FechaPedido, WEEK(MONDAY))  AS semana,
      ROUND(SUM(COALESCE(Litros, 0)), 2)     AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena grupo temple'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'GIN BOSQUE.*(500|750)')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY semana
    """
    q_so = f"""
    SELECT
      DATE_TRUNC(Fecha, WEEK(MONDAY)) AS semana,
      ROUND(SUM(Gin_Total), 2)        AS litros
    FROM `temple-bar-439715.curated_database.curated_gin`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
    GROUP BY semana
    """
    si_map = {str(r.semana): float(r.litros) for r in bq_dest.query(q_si).result(timeout=60)}
    so_map = {str(r.semana): float(r.litros) for r in bq_temple.query(q_so).result(timeout=60)}

    # Sell-in desde el corte: Ventas_Maestro está congelada desde 2026-07-01, se completa con Contabilium
    for (_local, wk), litros in _fetch_cbl_sellin_by_week_local(creds, r'GIN BOSQUE.*(500|750)').items():
        si_map[wk] = si_map.get(wk, 0.0) + litros

    all_weeks = sorted(set(list(si_map) + list(so_map)), reverse=True)[:weeks]

    result = []
    for i, wk in enumerate(all_weeks):
        si      = si_map.get(wk, 0.0)
        so      = so_map.get(wk, 0.0)
        prev_wk = all_weeks[i + 1] if i + 1 < len(all_weeks) else None
        si_prev = si_map.get(prev_wk) if prev_wk else None
        result.append({
            "w":       wk,
            "si":      si,
            "so":      so,
            "diff":    round(si - so, 2),
            "si_prev": si_prev,
            "alarm":   (si_prev is not None and so > si_prev),
        })
    return result


# ── Mapping sell-in NombreDeFantasia → nombre canónico ──────────────────────
# None = excluir del análisis cadena (no son locales de la cadena Grupo Temple)
_SI_ALIAS = {
    "Temple Craft Madero":              "Puerto Madero",
    "Temple Hollywood":                 "Hollywood",
    "MINIMARKET (Distri Rio Gallegos)": "Rio Gallegos",
    "Temple Barrio Chino":              "Barrio Chino + Monroe",
    "Temple Paseo La Plaza":            "Club Temple",
    "Temple Craft Soho":                "Soho",
    "Temple Monroe":                    "Barrio Chino + Monroe",
    "Temple Recoleta":                  "Recoleta",
    "Temple Santiago del Estero":       "Santiago del Estero",
    "Temple Craft Pilar":               "Pilar",
    "Temple Craft Salta":               "Salta",
    "Temple Comodoro":                  "Comodoro Rivadavia",
    "Temple Maschwitz":                 "Maschwitz",
    "Temple Cordoba":                   "Córdoba",
    "Temple Caminito":                  "Caminito",
    "Temple Palermo":                   "Casa Temple",
    "Temple Rio Gallegos":               "Rio Gallegos",
    "Temple Rosario":                    "Rosario 2",
    "Patagonia Santiago del estero":     None,   # cadena Patagonia, no Temple
    "Barra Patio de los Lecheros":      None,   # no es cadena Grupo Temple
    "Trenque Craft":                    None,   # no es cadena Grupo Temple
}

# ── Mapping sell-out Establecimiento → nombre canónico ──────────────────────
_SO_ALIAS = {
    "PUERTO MADERO":      "Puerto Madero",
    "HOLLYWOOD":          "Hollywood",
    "CLUB TEMPLE":        "Club Temple",
    "SOHO":               "Soho",
    "CASA TEMPLE":        "Casa Temple",
    "MASCHWITZ":          "Maschwitz",
    "RIO GALLEGOS":       "Rio Gallegos",
    "BARRIO CHINO":       "Barrio Chino + Monroe",
    "SALTA":              "Salta",
    "PILAR":              "Pilar",
    "MONROE":             "Barrio Chino + Monroe",
    "CORRIENTES":         "Corrientes",
    "RECOLETA":           "Recoleta",
    "SANTIAGO DEL ESTERO": "Santiago del Estero",
    "ROSARIO 2":          "Rosario 2",
    "COMODORO RIVADAVIA": "Comodoro Rivadavia",
    "TUCUMAN 3":          "Tucumán 3",
    "CAMINITO":           "Caminito",
    "GUEMES":             "Güemes",
    "PINAMAR":            "Pinamar",
}


# ── Feriado sell in/out — aliases NombreDeFantasia / Establecimiento ────────
_FER_SI_ALIAS: dict[str, str | None] = {
    "Temple Craft Madero":              "Puerto Madero",
    "Temple Hollywood":                 "Hollywood",
    "MINIMARKET (Distri Rio Gallegos)": "Rio Gallegos",
    "Temple Barrio Chino":              "Barrio Chino + Monroe",
    "Temple Paseo La Plaza":            "Club Temple",
    "Temple Craft Soho":                "Soho",
    "Temple Monroe":                    "Barrio Chino + Monroe",
    "Temple Recoleta":                  "Recoleta",
    "Temple Santiago del Estero":       "Santiago del Estero",
    "Temple Craft Pilar":               "Pilar",
    "Temple Craft Salta":               "Salta",
    "Temple Comodoro":                  "Comodoro Rivadavia",
    "Temple Maschwitz":                 "Maschwitz",
    "Temple Cordoba":                   "Córdoba",
    "Temple Caminito":                  "Caminito",
    "Temple Palermo":                   "Casa Temple",
    "Temple Rio Gallegos":               "Rio Gallegos",
    "Temple Rosario":                    "Rosario 2",
    "Patagonia Santiago del estero":     None,   # cadena Patagonia, no Temple
    "Barra Patio de los Lecheros":      None,
    "Trenque Craft":                    None,
}

_FER_SO_ALIAS: dict[str, str | None] = {
    "PUERTO MADERO":       "Puerto Madero",
    "HOLLYWOOD":           "Hollywood",
    "CLUB TEMPLE":         "Club Temple",
    "SOHO":                "Soho",
    "CASA TEMPLE":         "Casa Temple",
    "MASCHWITZ":           "Maschwitz",
    "RIO GALLEGOS":        "Rio Gallegos",
    "BARRIO CHINO":        "Barrio Chino + Monroe",
    "SALTA":               "Salta",
    "PILAR":               "Pilar",
    "MONROE":              "Barrio Chino + Monroe",
    "CORRIENTES":          "Corrientes",
    "RECOLETA":            "Recoleta",
    "SANTIAGO DEL ESTERO": "Santiago del Estero",
    "ROSARIO 2":           "Rosario 2",
    "COMODORO RIVADAVIA":  "Comodoro Rivadavia",
    "TUCUMAN 3":           "Tucumán 3",
    "CAMINITO":            "Caminito",
    "GUEMES":              "Güemes",
    "PINAMAR":             "Pinamar",
    "TAP ROOM":            None,  # no es local de cadena
}


def fetch_sellinout_fer_weekly(creds, weeks=12):
    """
    Sell-in/out semanal de Feriado para Cadena Grupo Temple.

    Sell-in : Ventas_Maestro_Con_Cluster_Final — cl='CADENA GRUPO TEMPLE',
              filtro REGEXP 'FERIADO' en Producto, campo Litros.
    Sell-out: curated_database.curated_feriado — campo Feriado_Total (litros).

    Alarma: sell-out[semana N] > sell-in[semana N-1]
    """
    from google.cloud import bigquery as _bq
    bq_dest   = _bq.Client(project="temple-brewery",    credentials=creds)
    bq_temple = _bq.Client(project="temple-bar-439715", credentials=creds)

    q_si = f"""
    SELECT
      DATE_TRUNC(FechaPedido, WEEK(MONDAY)) AS semana,
      ROUND(SUM(COALESCE(Litros, 0)), 2)    AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena grupo temple'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'FERIADO')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY semana
    """
    q_so = f"""
    SELECT
      DATE_TRUNC(Fecha, WEEK(MONDAY)) AS semana,
      ROUND(SUM(Feriado_Total), 2)    AS litros
    FROM `temple-bar-439715.curated_database.curated_feriado`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
    GROUP BY semana
    """
    si_map = {str(r.semana): float(r.litros) for r in bq_dest.query(q_si).result(timeout=60)}
    so_map = {str(r.semana): float(r.litros) for r in bq_temple.query(q_so).result(timeout=60)}

    # Sell-in desde el corte: Ventas_Maestro está congelada desde 2026-07-01, se completa con Contabilium
    for (_local, wk), litros in _fetch_cbl_sellin_by_week_local(creds, r'FERIADO').items():
        si_map[wk] = si_map.get(wk, 0.0) + litros

    all_weeks = sorted(set(list(si_map) + list(so_map)), reverse=True)[:weeks]

    result = []
    for i, wk in enumerate(all_weeks):
        si      = si_map.get(wk, 0.0)
        so      = so_map.get(wk, 0.0)
        prev_wk = all_weeks[i + 1] if i + 1 < len(all_weeks) else None
        si_prev = si_map.get(prev_wk) if prev_wk else None
        result.append({
            "w":       wk,
            "si":      si,
            "so":      so,
            "diff":    round(si - so, 2),
            "si_prev": si_prev,
            "alarm":   (si_prev is not None and so > si_prev),
        })
    return result


def fetch_sellinout_fer_local_weekly(creds, weeks=12):
    """Sell-in vs sell-out Feriado por local y por semana — pivot para vista Sem × Local."""
    from google.cloud import bigquery as _bq
    bq_dest   = _bq.Client(project="temple-brewery",    credentials=creds)
    bq_temple = _bq.Client(project="temple-bar-439715", credentials=creds)

    q_si = f"""
    SELECT
      DATE_TRUNC(FechaPedido, WEEK(MONDAY)) AS semana,
      TRIM(NombreDeFantasia)                AS local,
      ROUND(SUM(COALESCE(Litros, 0)), 2)    AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena grupo temple'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'FERIADO')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY semana, local
    """
    q_so = f"""
    SELECT
      DATE_TRUNC(Fecha, WEEK(MONDAY))   AS semana,
      TRIM(Establecimiento)             AS local,
      ROUND(SUM(Feriado_Total), 2)      AS litros
    FROM `temple-bar-439715.curated_database.curated_feriado`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
    GROUP BY semana, local
    """
    # Agrupar por (semana, local_canónico)
    si_data: dict[tuple, float] = {}
    for r in bq_dest.query(q_si).result(timeout=60):
        canon = _FER_SI_ALIAS.get(r.local, r.local)
        if canon is None:
            continue
        key = (str(r.semana), canon)
        si_data[key] = si_data.get(key, 0.0) + float(r.litros)

    # Sell-in desde el corte: Ventas_Maestro está congelada desde 2026-07-01, se completa con Contabilium
    for (local, wk), litros in _fetch_cbl_sellin_by_week_local(creds, r'FERIADO').items():
        key = (wk, local)
        si_data[key] = si_data.get(key, 0.0) + litros

    so_data: dict[tuple, float] = {}
    for r in bq_temple.query(q_so).result(timeout=60):
        canon = _FER_SO_ALIAS.get(r.local.upper() if r.local else "", r.local)
        if canon is None:
            continue
        key = (str(r.semana), canon)
        so_data[key] = so_data.get(key, 0.0) + float(r.litros)

    all_keys = set(si_data) | set(so_data)
    result = []
    for (wk, local) in sorted(all_keys, key=lambda x: (x[0], x[1]), reverse=False):
        si = round(si_data.get((wk, local), 0.0), 2)
        so = round(so_data.get((wk, local), 0.0), 2)
        if si == 0 and so == 0:
            continue
        result.append({"w": wk, "local": local, "si": si, "so": so})
    return result


# ── Feriado Patagonia — productos y aliases ──────────────────────────────────
_FER_PAT_SO_PRODUCTS = ['VERMU FERIADO', 'VERMU FERIADO CON TONICA', 'VERMU FERIADO CON POMELO']

# Alias Contabilium (post-migración, ver _CBL_SELLIN_CUTOFF) para completar la
# compra de Feriado de Cadena Patagonia — confirmado con Darwin 2026-08-11.
# También reusado para Bosque (misma cuenta de cliente, distinto producto) —
# ver fetch_sellinout_pat_weekly / fetch_sellinout_pat_local_weekly / _local_monthly.
_CBL_PAT_SI_ALIAS: dict[str, str | None] = {
    "PATAGONIA POSADAS BEER S.A.": "Mis - Posadas",
    # razón social compartida entre los dos locales de Puerto Iguazú.
    "PATAGONIA IGUAZU BEER":       "Mis - Puerto Iguazu (Ambos)",
    # local sin visibilidad de venta en curated_mix — se muestra la compra igual,
    # el sell-out queda en 0 a propósito (confirmado, no es bug).
    "PATAGONIA 24.7":              "24.7",
    # detectados 2026-08-12 comprando Bosque en Contabilium bajo cluster "Sin
    # Cluster" en lugar de "Cadena Patagonia" — el alias los rescata igual.
    "PATAGONIA LANUS":             "Ba - Lanus",
    "PATAGONIA NEUQUEN":           "Neuquen",
}

# Alias de sell-out (curated_mix) para que los dos Iguazú se sumen bajo el mismo
# canónico que usa _CBL_PAT_SI_ALIAS — cualquier local no listado usa .title().
_FER_PAT_SO_ALIAS: dict[str, str | None] = {
    "MIS - PUERTO IGUAZU":   "Mis - Puerto Iguazu (Ambos)",
    "MIS - PUERTO IGUAZU 2": "Mis - Puerto Iguazu (Ambos)",
}

_FER_PAT_SI_ALIAS: dict[str, str | None] = {
    "Patagonia Casa Tango":              "Casa Del Tango",
    "Patagonia Leloir":                  "Leloir",
    "Patagonia Bahia Blanca":            "Bahia Blanca",
    "Patagonia Mendoza":                 "Mendoza",
    "PATAGONIA Neuquen":                 "Neuquen",
    "CERVECER\u00cdA PATAGONIA Neuquen": "Neuquen",
    "Patagonia El Chalten":              "Chalten",
    "PATAGONIA USHUAIA":                 "Ushuaia",
    "PATAGONIA CALAFATE":                "Calafate",
    "PATAGONIA RIO GALLEGOS":            "Rio Gallegos",
    "Patagonia Armenia":                 "Ba - Plaza Armenia",
    "Patagonia Riobamba":                "Ba - Riobamba",
    "Patagonia Puerto Madero":           "Puerto Madero",
    "Refugio Patagonia Parana":          "Parana",
    "Patagonia Caril\u00f3":             "Carilo",
    "Refugio Patagonia Jujuy":           "Jujuy",
    "Patagonia Lanus":                   "Ba - Lanus",
    "CERVECERIA PATAGONIA (PINAMAR)":    "Pinamar",
    "PATAGONIA POSADAS BEER S.A.":       "Mis - Posadas",
    "PATAGONIA ROSARIO LA FLORIDA":      "Sfe - Rosario La Florida",
    "PATAGONIA LA PLATA 20 Y 50":        None,   # sin match en curated_mix
    "BARRIO DAMALE S.R.L.":              None,   # distribuidor
}


def fetch_sellinout_fer_pat_weekly(creds, weeks=12):
    """
    Sell-in/out semanal Feriado para Cadena Patagonia.
    Sell-in : Ventas_Maestro_Con_Cluster_Final — cl='cadena patagonia', FERIADO, campo Litros.
    Sell-out: patagonia-refugios.curated_database.curated_mix — VERMU FERIADO / FERIADITO,
              litros = SUM(tragos_total) * 0.24 (240ml por serve).
    """
    from google.cloud import bigquery as _bq
    bq_dest = _bq.Client(project="temple-brewery",     credentials=creds)
    bq_pat  = _bq.Client(project="patagonia-refugios", credentials=creds)

    q_si = f"""
    SELECT DATE_TRUNC(FechaPedido, WEEK(MONDAY)) AS semana,
           ROUND(SUM(COALESCE(Litros, 0)), 2)    AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena patagonia'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'FERIADO')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
    GROUP BY semana
    """
    _prods_sql = ', '.join(f"'{p}'" for p in _FER_PAT_SO_PRODUCTS)
    q_so = f"""
    SELECT DATE_TRUNC(fecha, WEEK(MONDAY))            AS semana,
           ROUND(SUM(COALESCE(tragos_total, 0)) * 0.12, 2) AS litros
    FROM `patagonia-refugios.curated_database.curated_mix`
    WHERE fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND UPPER(TRIM(producto)) IN ({_prods_sql})
      AND categoria = 'VERMUTH'
    GROUP BY semana
    """
    si_map = {str(r.semana): float(r.litros) for r in bq_dest.query(q_si).result(timeout=60)}
    so_map = {str(r.semana): float(r.litros) for r in bq_pat.query(q_so).result(timeout=60)}

    # Ventas_Maestro está congelada (ver lesson_sin-compra-datos-reales) — se completa
    # la compra reciente desde Contabilium, igual que para Cadena Grupo Temple.
    for (_local, wk), litros in _fetch_cbl_sellin_by_week_local(
            creds, r'FERIADO', alias_map=_CBL_PAT_SI_ALIAS, cluster_name='Cadena Patagonia').items():
        si_map[wk] = si_map.get(wk, 0.0) + litros

    all_weeks = sorted(set(list(si_map) + list(so_map)), reverse=True)[:weeks]
    result = []
    for i, wk in enumerate(all_weeks):
        si      = si_map.get(wk, 0.0)
        so      = so_map.get(wk, 0.0)
        prev_wk = all_weeks[i + 1] if i + 1 < len(all_weeks) else None
        si_prev = si_map.get(prev_wk) if prev_wk else None
        result.append({"w": wk, "si": si, "so": so, "diff": round(si - so, 2),
                        "si_prev": si_prev, "alarm": (si_prev is not None and so > si_prev)})
    return result


def fetch_sellinout_fer_pat_local_weekly(creds, weeks=12):
    """Sell-in vs sell-out Feriado Patagonia por local × semana."""
    from google.cloud import bigquery as _bq
    bq_dest = _bq.Client(project="temple-brewery",     credentials=creds)
    bq_pat  = _bq.Client(project="patagonia-refugios", credentials=creds)

    q_si = f"""
    SELECT DATE_TRUNC(FechaPedido, WEEK(MONDAY)) AS semana,
           TRIM(NombreDeFantasia)                AS local,
           ROUND(SUM(COALESCE(Litros, 0)), 2)   AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena patagonia'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'FERIADO')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
    GROUP BY semana, local
    """
    _prods_sql = ', '.join(f"'{p}'" for p in _FER_PAT_SO_PRODUCTS)
    q_so = f"""
    SELECT DATE_TRUNC(fecha, WEEK(MONDAY))                   AS semana,
           TRIM(establecimiento)                              AS local,
           ROUND(SUM(COALESCE(tragos_total, 0)) * 0.12, 2)  AS litros
    FROM `patagonia-refugios.curated_database.curated_mix`
    WHERE fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND UPPER(TRIM(producto)) IN ({_prods_sql})
      AND categoria = 'VERMUTH'
    GROUP BY semana, local
    """
    si_data: dict[tuple, float] = {}
    for r in bq_dest.query(q_si).result(timeout=60):
        canon = _FER_PAT_SI_ALIAS.get(r.local, r.local)
        if canon is None:
            continue
        key = (str(r.semana), canon)
        si_data[key] = si_data.get(key, 0.0) + float(r.litros)

    # Completar compra reciente desde Contabilium (Ventas_Maestro congelada, ver
    # lesson_sin-compra-datos-reales) — mismo mecanismo que Cadena Grupo Temple.
    for (local, wk), litros in _fetch_cbl_sellin_by_week_local(
            creds, r'FERIADO', alias_map=_CBL_PAT_SI_ALIAS, cluster_name='Cadena Patagonia').items():
        key = (wk, local)
        si_data[key] = si_data.get(key, 0.0) + litros

    so_data: dict[tuple, float] = {}
    for r in bq_pat.query(q_so).result(timeout=60):
        raw = r.local or ""
        canon = _FER_PAT_SO_ALIAS.get(raw, raw.title())
        if canon is None:
            continue
        key = (str(r.semana), canon)
        so_data[key] = so_data.get(key, 0.0) + float(r.litros)

    all_keys = set(si_data) | set(so_data)
    result = []
    for (wk, local) in sorted(all_keys):
        si = round(si_data.get((wk, local), 0.0), 2)
        so = round(so_data.get((wk, local), 0.0), 2)
        if si == 0 and so == 0:
            continue
        result.append({"w": wk, "local": local, "si": si, "so": so})
    return result


def fetch_sellinout_fer_pat_local_monthly(creds, months=3):
    """Compra vs venta Feriado Cadena Patagonia por local y MES CALENDARIO — usado
    por los filtros Mes Actual / Mes Anterior cuando la vista está en Cadena
    Patagonia (mismo objetivo que fetch_sellinout_local_monthly para Grupo Temple,
    pero con las queries dedicadas de Patagonia — el sell-out no es un SUM simple,
    sino tragos_total*0.12 filtrado por producto/categoria)."""
    from google.cloud import bigquery as _bq
    bq_dest = _bq.Client(project="temple-brewery",     credentials=creds)
    bq_pat  = _bq.Client(project="patagonia-refugios", credentials=creds)

    q_si = f"""
    SELECT FORMAT_DATE('%Y-%m', FechaPedido) AS mes,
           TRIM(NombreDeFantasia)            AS local,
           ROUND(SUM(COALESCE(Litros, 0)), 2) AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena patagonia'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'FERIADO')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {months} MONTH)
    GROUP BY mes, local
    """
    _prods_sql = ', '.join(f"'{p}'" for p in _FER_PAT_SO_PRODUCTS)
    q_so = f"""
    SELECT FORMAT_DATE('%Y-%m', fecha)                     AS mes,
           TRIM(establecimiento)                             AS local,
           ROUND(SUM(COALESCE(tragos_total, 0)) * 0.12, 2) AS litros
    FROM `patagonia-refugios.curated_database.curated_mix`
    WHERE fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {months} MONTH)
      AND UPPER(TRIM(producto)) IN ({_prods_sql})
      AND categoria = 'VERMUTH'
    GROUP BY mes, local
    """

    si_map: dict[tuple, float] = {}
    for r in bq_dest.query(q_si).result(timeout=60):
        canon = _FER_PAT_SI_ALIAS.get(r.local, r.local)
        if canon is None:
            continue
        key = (canon, r.mes)
        si_map[key] = si_map.get(key, 0.0) + float(r.litros)

    for (local, mes), litros in _fetch_cbl_sellin_by_month_local(
            creds, r'FERIADO', alias_map=_CBL_PAT_SI_ALIAS, cluster_name='Cadena Patagonia').items():
        key = (local, mes)
        si_map[key] = si_map.get(key, 0.0) + litros

    so_map: dict[tuple, float] = {}
    for r in bq_pat.query(q_so).result(timeout=60):
        raw = r.local or ""
        canon = _FER_PAT_SO_ALIAS.get(raw, raw.title())
        if canon is None:
            continue
        key = (canon, r.mes)
        so_map[key] = so_map.get(key, 0.0) + float(r.litros)

    result = []
    for (local, mes) in sorted(set(si_map) | set(so_map)):
        si = round(si_map.get((local, mes), 0.0), 2)
        so = round(so_map.get((local, mes), 0.0), 2)
        if si == 0 and so == 0:
            continue
        result.append({"local": local, "mo": mes, "si": si, "so": so})
    return result


# ── Productos Patagonia que usan 50ml de Gin Bosque en receta ───────────────
_PAT_SO_PRODUCTS = [
    'GIN TONIC - BOSQUE GIN',
    'BOTELLA GIN TONIC - (BOSQUE GIN)',
    'GIN TONIC LIMON',
    'GIN TONIC LIMA',
    'GIN TONIC POMELO',
    'GIN TONIC NARANJA',
    'GIN PEPINO',
    'GIN FRUTOS',
    'GIN MARACUYA',
    'GIN PEPINO LIMON',
    'GIN BOTELLA MEDIDA',
    'DRAGON GIN',
]
# Alias NombreDeFantasia → nombre canónico (None = excluir).
_PAT_SI_ALIAS: dict[str, str | None] = {
    "Patagonia Casa Tango":      "Casa Del Tango",
    "Patagonia Lanus":           "Ba - Lanus",
    "Patagonia Leloir":          "Leloir",
    "Patagonia Mendoza":         "Mendoza",
    "Refugio Patagonia Parana":  "Parana",
}
# Alias Establecimiento → nombre canónico. Completar con las sucursales de Patagonia.
_PAT_SO_ALIAS: dict[str, str | None] = {
    "MIS - PUERTO IGUAZU":   "Mis - Puerto Iguazu (Ambos)",
    "MIS - PUERTO IGUAZU 2": "Mis - Puerto Iguazu (Ambos)",
}


def fetch_sellinout_by_local(creds, weeks=12):
    """Sell-in (NombreDeFantasia) vs sell-out (Establecimiento) totales por local, últimas N semanas."""
    from google.cloud import bigquery as _bq
    bq_dest   = _bq.Client(project="temple-brewery",    credentials=creds)
    bq_temple = _bq.Client(project="temple-bar-439715", credentials=creds)

    q_si = f"""
    SELECT
      TRIM(NombreDeFantasia)             AS local,
      ROUND(SUM(COALESCE(Litros, 0)), 2) AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena grupo temple'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'GIN BOSQUE.*(500|750)')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks} WEEK)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY local
    """

    q_so = f"""
    SELECT
      TRIM(Establecimiento)    AS local,
      ROUND(SUM(Gin_Total), 2) AS litros
    FROM `temple-bar-439715.curated_database.curated_gin`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks} WEEK)
    GROUP BY local
    """

    si_raw = {r.local: float(r.litros) for r in bq_dest.query(q_si).result(timeout=60)}
    so_raw = {r.local: float(r.litros) for r in bq_temple.query(q_so).result(timeout=60)}

    # Aplicar alias sell-in (None = excluir)
    si_map: dict[str, float] = {}
    for name, litros in si_raw.items():
        canon = _SI_ALIAS.get(name, name)
        if canon is None:
            continue
        si_map[canon] = si_map.get(canon, 0.0) + litros

    # Sell-in desde el corte: Ventas_Maestro está congelada desde 2026-07-01, se completa con Contabilium
    for (local, _wk), litros in _fetch_cbl_sellin_by_week_local(creds, r'GIN BOSQUE.*(500|750)').items():
        si_map[local] = si_map.get(local, 0.0) + litros

    # Aplicar alias sell-out
    so_map: dict[str, float] = {}
    for name, litros in so_raw.items():
        canon = _SO_ALIAS.get(name, name.title())
        so_map[canon] = so_map.get(canon, 0.0) + litros

    all_locals = sorted(set(list(si_map) + list(so_map)))
    result = []
    for local in all_locals:
        si = si_map.get(local, 0.0)
        so = so_map.get(local, 0.0)
        result.append({"local": local, "si": si, "so": so, "diff": round(si - so, 2)})

    result.sort(key=lambda x: x["so"], reverse=True)
    return result


def fetch_sellinout_local_weekly(creds, weeks=12):
    """Sell-in vs sell-out por local y por semana — pivot para vista Sem × Local."""
    from google.cloud import bigquery as _bq
    bq_dest   = _bq.Client(project="temple-brewery",    credentials=creds)
    bq_temple = _bq.Client(project="temple-bar-439715", credentials=creds)

    q_si = f"""
    SELECT
      DATE_TRUNC(FechaPedido, WEEK(MONDAY)) AS semana,
      TRIM(NombreDeFantasia)                AS local,
      ROUND(SUM(COALESCE(Litros, 0)), 2)   AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena grupo temple'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'GIN BOSQUE.*(500|750)')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY semana, local
    """

    q_so = f"""
    SELECT
      DATE_TRUNC(Fecha, WEEK(MONDAY)) AS semana,
      TRIM(Establecimiento)           AS local,
      ROUND(SUM(Gin_Total), 2)        AS litros
    FROM `temple-bar-439715.curated_database.curated_gin`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
    GROUP BY semana, local
    """

    # Build maps: (canon_local, week_str) -> litros
    si_map: dict[tuple, float] = {}
    for r in bq_dest.query(q_si).result(timeout=60):
        canon = _SI_ALIAS.get(r.local, r.local)
        if canon is None:
            continue
        key = (canon, str(r.semana))
        si_map[key] = si_map.get(key, 0.0) + float(r.litros)

    # Sell-in desde el corte: Ventas_Maestro está congelada desde 2026-07-01, se completa con Contabilium
    for (local, wk), litros in _fetch_cbl_sellin_by_week_local(creds, r'GIN BOSQUE.*(500|750)').items():
        key = (local, wk)
        si_map[key] = si_map.get(key, 0.0) + litros

    so_map: dict[tuple, float] = {}
    for r in bq_temple.query(q_so).result(timeout=60):
        canon = _SO_ALIAS.get(r.local, r.local.title())
        key = (canon, str(r.semana))
        so_map[key] = so_map.get(key, 0.0) + float(r.litros)

    all_weeks  = sorted({k[1] for k in list(si_map) + list(so_map)}, reverse=True)[:weeks]
    all_locals = sorted({k[0] for k in list(si_map) + list(so_map)})

    result = []
    for local in all_locals:
        for wk in all_weeks:
            si = si_map.get((local, wk), 0.0)
            so = so_map.get((local, wk), 0.0)
            if si > 0 or so > 0:
                result.append({"local": local, "w": wk, "si": si, "so": so})
    return result


def fetch_sellinout_pat_weekly(creds, weeks=12):
    """
    Sell-in/out semanal para Cadena Patagonia.
    Sell-in : Ventas_Maestro_Con_Cluster_Final — cl='cadena patagonia', GIN BOSQUE 500+750ml, campo Litros.
    Sell-out: patagonia-refugios.curated_database.curated_mix — 12 productos con receta 50ml gin Bosque,
              litros = SUM(Cantidad) * 0.05.
    Alarma: sell-out[semana N] > sell-in[semana N-1].
    """
    from google.cloud import bigquery as _bq
    bq_dest = _bq.Client(project="temple-brewery",     credentials=creds)
    bq_pat  = _bq.Client(project="patagonia-refugios", credentials=creds)

    q_si = f"""
    SELECT
      DATE_TRUNC(FechaPedido, WEEK(MONDAY)) AS semana,
      ROUND(SUM(COALESCE(Litros, 0)), 2)   AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena patagonia'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'GIN BOSQUE.*(500|750)')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY semana
    """

    _prods_sql = ', '.join(f"'{p}'" for p in _PAT_SO_PRODUCTS)
    q_so = f"""
    SELECT
      DATE_TRUNC(Fecha, WEEK(MONDAY))     AS semana,
      ROUND(SUM(Cantidad) * 0.05, 2)     AS litros
    FROM `patagonia-refugios.curated_database.curated_mix`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND UPPER(TRIM(Producto)) IN ({_prods_sql})
    GROUP BY semana
    """

    si_map = {str(r.semana): float(r.litros) for r in bq_dest.query(q_si).result(timeout=60)}
    so_map = {str(r.semana): float(r.litros) for r in bq_pat.query(q_so).result(timeout=60)}

    # Sell-in desde el corte: Ventas_Maestro está congelada desde 2026-06-25 para
    # Cadena Patagonia, se completa con Contabilium (mismo patrón que Grupo Temple).
    for (_local, wk), litros in _fetch_cbl_sellin_by_week_local(
            creds, r'GIN BOSQUE.*(500|750)', alias_map=_CBL_PAT_SI_ALIAS, cluster_name='Cadena Patagonia').items():
        si_map[wk] = si_map.get(wk, 0.0) + litros

    all_weeks = sorted(set(list(si_map) + list(so_map)), reverse=True)[:weeks]
    result = []
    for i, wk in enumerate(all_weeks):
        si      = si_map.get(wk, 0.0)
        so      = so_map.get(wk, 0.0)
        prev_wk = all_weeks[i + 1] if i + 1 < len(all_weeks) else None
        si_prev = si_map.get(prev_wk) if prev_wk else None
        result.append({
            "w":       wk,
            "si":      si,
            "so":      so,
            "diff":    round(si - so, 2),
            "si_prev": si_prev,
            "alarm":   (si_prev is not None and so > si_prev),
        })
    return result


def fetch_sellinout_pat_local_weekly(creds, weeks=12):
    """Sell-in vs sell-out Patagonia por local y por semana — pivot para vista Sem × Local."""
    from google.cloud import bigquery as _bq
    bq_dest = _bq.Client(project="temple-brewery",     credentials=creds)
    bq_pat  = _bq.Client(project="patagonia-refugios", credentials=creds)

    q_si = f"""
    SELECT
      DATE_TRUNC(FechaPedido, WEEK(MONDAY)) AS semana,
      TRIM(NombreDeFantasia)                AS local,
      ROUND(SUM(COALESCE(Litros, 0)), 2)   AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena patagonia'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'GIN BOSQUE.*(500|750)')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY semana, local
    """

    _prods_sql = ', '.join(f"'{p}'" for p in _PAT_SO_PRODUCTS)
    q_so = f"""
    SELECT
      DATE_TRUNC(Fecha, WEEK(MONDAY))     AS semana,
      TRIM(Establecimiento)               AS local,
      ROUND(SUM(Cantidad) * 0.05, 2)     AS litros
    FROM `patagonia-refugios.curated_database.curated_mix`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {weeks + 2} WEEK)
      AND UPPER(TRIM(Producto)) IN ({_prods_sql})
    GROUP BY semana, local
    """

    si_map: dict[tuple, float] = {}
    for r in bq_dest.query(q_si).result(timeout=60):
        canon = _PAT_SI_ALIAS.get(r.local, r.local)
        if canon is None:
            continue
        key = (canon, str(r.semana))
        si_map[key] = si_map.get(key, 0.0) + float(r.litros)

    # Sell-in desde el corte: Ventas_Maestro está congelada desde 2026-06-25 para
    # Cadena Patagonia, se completa con Contabilium (mismo patrón que Grupo Temple).
    for (local, wk), litros in _fetch_cbl_sellin_by_week_local(
            creds, r'GIN BOSQUE.*(500|750)', alias_map=_CBL_PAT_SI_ALIAS, cluster_name='Cadena Patagonia').items():
        key = (local, wk)
        si_map[key] = si_map.get(key, 0.0) + litros

    so_map: dict[tuple, float] = {}
    for r in bq_pat.query(q_so).result(timeout=60):
        canon = _PAT_SO_ALIAS.get(r.local, r.local.title() if r.local else r.local)
        key = (canon, str(r.semana))
        so_map[key] = so_map.get(key, 0.0) + float(r.litros)

    all_weeks  = sorted({k[1] for k in list(si_map) + list(so_map)}, reverse=True)[:weeks]
    all_locals = sorted({k[0] for k in list(si_map) + list(so_map)})

    result = []
    for local in all_locals:
        for wk in all_weeks:
            si = si_map.get((local, wk), 0.0)
            so = so_map.get((local, wk), 0.0)
            if si > 0 or so > 0:
                result.append({"local": local, "w": wk, "si": si, "so": so})
    return result


def fetch_sellinout_pat_local_monthly(creds, months=3):
    """Compra vs venta Bosque Cadena Patagonia por local y MES CALENDARIO — usado
    por los filtros Mes Actual / Mes Anterior cuando la vista está en Cadena
    Patagonia (mismo objetivo que fetch_sellinout_local_monthly para Grupo Temple
    y fetch_sellinout_fer_pat_local_monthly para Feriado, evita el sesgo del
    bucket semanal en los bordes del mes)."""
    from google.cloud import bigquery as _bq
    bq_dest = _bq.Client(project="temple-brewery",     credentials=creds)
    bq_pat  = _bq.Client(project="patagonia-refugios", credentials=creds)

    q_si = f"""
    SELECT FORMAT_DATE('%Y-%m', FechaPedido) AS mes,
           TRIM(NombreDeFantasia)            AS local,
           ROUND(SUM(COALESCE(Litros, 0)), 2) AS litros
    FROM `temple-brewery.Destileria.Ventas_Maestro_Con_Cluster_Final`
    WHERE LOWER(TRIM(Clusterizacion)) = 'cadena patagonia'
      AND REGEXP_CONTAINS(UPPER(TRIM(Producto)), r'GIN BOSQUE.*(500|750)')
      AND FechaPedido >= DATE_SUB(CURRENT_DATE(), INTERVAL {months} MONTH)
      AND FechaPedido < '{_CBL_SELLIN_CUTOFF}'
    GROUP BY mes, local
    """
    _prods_sql = ', '.join(f"'{p}'" for p in _PAT_SO_PRODUCTS)
    q_so = f"""
    SELECT FORMAT_DATE('%Y-%m', Fecha)  AS mes,
           TRIM(Establecimiento)        AS local,
           ROUND(SUM(Cantidad) * 0.05, 2) AS litros
    FROM `patagonia-refugios.curated_database.curated_mix`
    WHERE Fecha >= DATE_SUB(CURRENT_DATE(), INTERVAL {months} MONTH)
      AND UPPER(TRIM(Producto)) IN ({_prods_sql})
    GROUP BY mes, local
    """

    si_map: dict[tuple, float] = {}
    for r in bq_dest.query(q_si).result(timeout=60):
        canon = _PAT_SI_ALIAS.get(r.local, r.local)
        if canon is None:
            continue
        key = (canon, r.mes)
        si_map[key] = si_map.get(key, 0.0) + float(r.litros)

    # Sell-in desde el corte: Ventas_Maestro está congelada desde 2026-06-25 para
    # Cadena Patagonia, se completa con Contabilium (mismo patrón que Grupo Temple).
    for (local, mes), litros in _fetch_cbl_sellin_by_month_local(
            creds, r'GIN BOSQUE.*(500|750)', alias_map=_CBL_PAT_SI_ALIAS, cluster_name='Cadena Patagonia').items():
        key = (local, mes)
        si_map[key] = si_map.get(key, 0.0) + litros

    so_map: dict[tuple, float] = {}
    for r in bq_pat.query(q_so).result(timeout=60):
        raw = r.local or ""
        canon = _PAT_SO_ALIAS.get(raw, raw.title())
        if canon is None:
            continue
        key = (canon, r.mes)
        so_map[key] = so_map.get(key, 0.0) + float(r.litros)

    result = []
    for (local, mes) in sorted(set(si_map) | set(so_map)):
        si = round(si_map.get((local, mes), 0.0), 2)
        so = round(so_map.get((local, mes), 0.0), 2)
        if si == 0 and so == 0:
            continue
        result.append({"local": local, "mo": mes, "si": si, "so": so})
    return result


if __name__ == "__main__":
    main()
